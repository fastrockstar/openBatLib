import tools
import numpy as np
import scipy.io as sio
import numba as nb
from numba import types
from numba.typed import Dict
from numba import njit
import pandas as pd
import time
import datetime
import csv
from openpyxl import load_workbook
from pyModbusTCP.client import ModbusClient
from pyModbusTCP import utils


class BatModDC(object):
    """Performance Simulation Class for DC-coupled PV-Battery systems

    :param object: object
    :type object: object
    """
    _version = 0.1

    def __init__(self, parameter, d, ppv, pl, Pr, Prpv, Ppv, ppv2ac, Ppv2ac_out, dt):
        self.parameter = parameter
        self.d = d
        self.ppv = ppv
        self.pl = pl
        self.Pr = Pr
        self.Prpv = Prpv
        self.Ppv = Ppv
        self.ppv2ac = ppv2ac
        self.Ppv2ac_out = Ppv2ac_out
        self.Ppv2ac_out0 = 0
        self.Ppv2bat_in0 = 0

        self.dt = dt
        # Initialization and preallocation
        self.Pbat = np.zeros_like(self.ppv)  # DC power of the battery in W
        self.soc = np.zeros_like(self.ppv)  # State of charge of the battery
        # Input power of the PV2BAT conversion pathway in W
        self.Ppv2bat_in = np.zeros_like(self.ppv)
        # Output power of the BAT2AC conversion pathway in W
        self.Pbat2ac_out = np.zeros_like(self.ppv)
        self.Pbat2ac_out0 = 0
        # AC power of the PV-battery system in W
        self.Ppvbs = np.zeros_like(self.ppv)

        self.th = 0  # Start threshold for the recharging of the battery
        self.soc0 = 0  # State of charge of the battery in the first time step

        # Additional power consumption of other system components (e.g. AC power meter) in W
        self.Pperi = np.ones(self.ppv.size) * self.parameter['P_PERI_AC']

        self.simulation()

        self.bat_mod_res()

    def simulation(self, pvmod=True):

        self.Ppv2ac_out, self.Ppv2bat_in, self.Ppv2bat_in0, self.Pbat2ac_out, self.Pbat2ac_out0, self.Ppvbs, self.Pbat, self.soc, self.soc0 = BatMod_DC(self.d, self.dt, self.soc0, self.soc, self.Pr, self.Prpv,  self.Ppv, self.Ppv2bat_in0, self.Ppv2bat_in, self.Pbat2ac_out0, self.Pbat2ac_out, self.Ppv2ac_out0, self.Ppv2ac_out, self.Ppvbs, self.Pbat)
        
        # Define missing parameters
        self.Ppv2ac = self.Ppv2ac_out  # AC output power of the PV2AC conversion pathway
        self.Ppv2bat = self.Ppv2bat_in  # DC input power of the PV2BAT conversion pathway

    def bat_mod_res(self):
        self.E = bat_res_mod(self.parameter, self.pl, self.Ppv, self.Pbat,
                                   self.dt, self.Ppv2ac, self.Ppv2bat, self.Ppvbs, self.Pperi)

    def get_E(self):
        return self.E

    def get_soc(self):
        return self.soc

    def get_Pbat(self):
        return self.Pbat


class BatModAC(object):
    """Performance Simulation Class for AC-coupled PV-Battery systems

    :param object: object
    :type object: object
    """
    _version = '0.1'

    def __init__(self, parameter, d, ppv, pl, Pr, Ppv, Ppvs, Pperi, dt):
        self.parameter = parameter
        self.d = d
        self.ppv = ppv
        self.pl = pl
        self.Pr = Pr
        self.Ppv = Ppv
        self.Ppvs = Ppvs
        self.Pperi = Pperi
        self.dt = dt

        # Initialization and preallocation
        self.Pbat = np.zeros_like(self.ppv)  # DC power of the battery in W
        self.Pbs = np.zeros_like(ppv)
        self.soc = np.zeros_like(self.ppv)  # State of charge of the battery
        self.th = 0  # Start threshold for the recharging of the battery
        self.soc0 = 0  # State of charge of the battery in the first time step
        self.Pbs0 = 0

        self.simulation()

        self.bat_mod_res()

    def simulation(self, pvmod=True):
        # PerModAC: Performance Simulation Model for AC-coupled PV-Battery Systems
        '''
        TODO: Preallocation verschieben
              Pr anpassen
        '''

        # 3.3 Simulation of the battery system   
        #start = time.process_time()     
        self.Pbat, self.Pbs, self.soc, self.soc0, self.Pbs0 = BatMod_AC(
            self.d, self.dt, self.soc0, self.soc, self.Pr, self.Pbs0, self.Pbs, self.Pbat)
        #print(time.process_time()-start)

        
    def bat_mod_res(self):
        self.E = bat_res_mod(
            self.parameter, self.pl, self.Ppv, self.Pbat, self.dt, self.Ppvs, self.Pbs, self.Pperi)

    def get_E(self):
        return self.E

    def get_soc(self):
        '''
        idx = pd.date_range(start='00:00:00', periods=len(self.soc), freq='S')
        _soc = pd.Series(self.soc, index=idx)
        soc_1h = _soc.resample('1h').sum()
        soc_1h = soc_1h.to_numpy()
        '''
        return self.soc

    def get_Pbat(self):
        return self.Pbat

    def get_Pbs(self):
        return self.Pbs


class BatModPV(object):
    """Performance Simulation Class for PV-coupled PV-Battery systems

    :param object: object
    :type object: object
    """
    _version = '0.1'

    def __init__(self, parameter, d, ppv, pl, Pac, Ppv, Pperi, dt):
        self.parameter = parameter
        self.d = d
        self.ppv = ppv
        self.pl = pl
        self.Pac = Pac
        self.Ppv = Ppv
        self.Pperi = Pperi
        self.dt = dt

        # Initialization and preallocation
        self.Pbat = np.zeros_like(self.ppv)  # DC power of the battery in W
        self.soc = np.zeros_like(self.ppv)  # State of charge of the battery
        # Output power of the PV2AC conversion pathway in W
        self.Ppv2ac_out = np.zeros_like(self.ppv)
        # Input power of the PV2BAT conversion pathway in W
        self.Ppv2bat_in = np.zeros_like(self.ppv)
        self.Ppv2bat_in0 = 0
        # Output power of the BAT2PV conversion pathway in W
        self.Pbat2pv_out = np.zeros_like(self.ppv)
        self.Pbat2pv_out0 = 0
        # AC power of the PV-battery system in W
        self.Ppvbs = np.zeros_like(self.ppv)

        self.simulation()

        self.bat_mod_res()

    def simulation(self, pvmod=True):

        self.th = 0  # Start threshold for the recharging of the battery
        self.soc0 = 0  # Initial state of charge of the battery in the first time step

        # Simulation of the battery system
        #start = time.process_time()
        self.soc, self.soc0, self.Ppv, self.Ppvbs, self.Pbat, self.Ppv2ac_out, self.Pbat2pv_out, self.Ppv2bat_in = BatMod_PV(self.d, self.dt, self.soc0, self.soc, self.Ppv, self.Pac, self.Ppv2bat_in0, self.Ppv2bat_in, self.Ppv2ac_out, self.Pbat2pv_out0, self.Pbat2pv_out, self.Ppvbs, self.Pbat)
        #print(time.process_time()-start)
        # Define missing parameters
        self.Ppv2ac = self.Ppv2ac_out  # AC output power of the PV2AC conversion pathway
        self.Ppv2bat = self.Ppv2bat_in  # DC input power of the PV2BAT conversion pathway

    def bat_mod_res(self):
        self.E = bat_res_mod(self.parameter, self.pl, self.Ppv, self.Pbat, self.dt, self.Ppv2ac, self.Ppv2bat, self.Ppvbs, self.Pperi)

    def get_E(self):
        return self.E

    def get_soc(self):
        '''
        idx = pd.date_range(start='00:00:00', periods=len(self.soc), freq='S')
        _soc = pd.Series(self.soc, index=idx)
        soc_1h = _soc.resample('1h').sum()
        soc_1h = soc_1h.to_numpy()
        '''
        return self.soc

    def get_Pbat(self):
        return self.Pbat


class ModBus(object):
    """Establishes connection to a battery system via ModBus protocol

    :param object: object
    :type object: object
    """

    def __init__(self, host, port, unit_id, input_vals, dt, fname):
        self.host = host
        self.port = port
        self.unit_id = unit_id
        self.dt = dt
        self.input_vals = input_vals
        self.fname = fname
        self.open_connection()
        self.create_csv_file()

        self.start_loop()

    def open_connection(self):

        # Open ModBus connection
        try:
            self.c = ModbusClient(host=self.host, port=self.port,
                                  unit_id=self.unit_id, auto_open=True, auto_close=True)
        except ValueError:
            print("Error with host: {}, port: {} or unit-ID: {} params".format(
                self.host, self.port, self.unit_id))
        # Arrray for the setting values

    def start_loop(self):
        # Transform the array to fit the 1 minute time duration
        #self.set_vals = np.repeat(self.input_vals, self.dt * 60)

        i = 0
        idx = pd.date_range(start=datetime.datetime.now(),
                            periods=(self.input_vals.size), freq='S')
        while i < len(idx):
            if datetime.datetime.now().second == idx[i].second:
                # Set chrging value
                self.set_val = int(self.input_vals[i])
                if self.set_val < 0:
                    # Write negative value to battery charge power (AC) setpoint register
                    self.c.write_single_register(1024, self.set_val & 0xFFFF)
                    # Log writing time
                    self.set_time = datetime.datetime.now()
                else:
                    # Write positive value to battery charge power (AC) setpoint to register
                    self.c.write_single_register(1024, self.set_val)
                    # Log writing time
                    self.set_time = datetime.datetime.now()

                try:
                    # Read total AC power value from register
                    _P_ac = self.c.read_holding_registers(172, 2)
                    self.read_time_P_ac = datetime.datetime.now()
                except:
                    print('Could not read register 172!')

                try:
                    # Read actual battery charge/discharge power value from register
                    _P_bat = self.c.read_holding_registers(582, 1)
                    self.read_time_P_bat = datetime.datetime.now()
                except:
                    print('Could not read register 582!')

                # Load content of two registers into a single float value
                zregs = utils.word_list_to_long(_P_ac, big_endian=False)
                # Decode and store float value of the AC-power
                self.P_ac = utils.decode_ieee(*zregs)
                # Store the DC charging power
                self.P_bat = np.int16(*_P_bat)
                # Read actual soc
                self.soc0 = self.read_soc(210)

                try:
                    # Save the values to a csv file
                    self.save_to_csv()
                except:
                    print('Could not save to csv!')

                i += 1

    def read_soc(self, reg):
        # Load the actual state fo charge of the battery
        regs = self.c.read_holding_registers(reg, 2)
        # Load content of two registers into a single float value
        zregs = utils.word_list_to_long(regs, big_endian=False)

        return utils.decode_ieee(*zregs)

    def create_csv_file(self):
        # Create a new csv-file
        with open(self.fname, 'w') as f:
            writer = csv.writer(f, dialect='excel')
            writer.writerow(['set_time',
                             'read_time_P_ac',
                             'read_time_P_bat',
                             'soc',
                             'set_value',
                             'P_ac',
                             'P_bat'])

    def save_to_csv(self):
        # Save the read values to a csv file
        with open(self.fname, "a") as f:
            wr = csv.writer(f, dialect='excel')
            wr.writerow([self.set_time, self.read_time_P_ac, self.read_time_P_bat,
                         self.soc0, self.set_val, self.P_ac, self.P_bat])


def max_self_consumption(parameter, ppv, pl, pvmod=True):
    """Function for maximizing self consumption

    :param object: object
    :ppv object: object
    :pl object: object
    """
    # Maximize self consumption for AC-coupled systems
    if parameter['Top'] == 'AC':

        # DC power output of the PV generator
        if pvmod:  # ppv: Normalized DC power output of the PV generator in kW/kWp
            Ppv = np.minimum(
                ppv * parameter['P_PV'], parameter['P_PV2AC_in']) * 1000
        else:  # ppv: DC power output of the PV generator in W
            Ppv = np.minimum(ppv, parameter['P_PV2AC_in'] * 1000)

        # Normalized input power of the PV inverter
        ppvinvin = Ppv / parameter['P_PV2AC_in'] / 1000

        # AC power output of the PV inverter taking into account the conversion losses and maximum
        # output power of the PV inverter
        Ppvs = np.minimum(np.maximum(0, Ppv-(parameter['PV2AC_a_in'] * ppvinvin * ppvinvin + parameter['PV2AC_b_in'] * ppvinvin + parameter['PV2AC_c_in'])), parameter['P_PV2AC_out'] * 1000)

        # 3.2 Residual power

        # Additional power consumption of other system components (e.g. AC power meter) in W
        Pperi = np.ones_like(ppv) * parameter['P_PERI_AC']

        # Adding the standby consumption of the PV inverter in times without any AC power output of the PV system
        # to the additional power consumption
        Pperi[Ppvs == 0] += parameter['P_PVINV_AC']

        # Residual power
        Pr = Ppvs - pl - Pperi

        return Pr, Ppv, Ppvs, Pperi

    # Maximize self consumption for DC-coupled systems
    elif parameter['Top'] == 'DC':
        # Initialization and preallocation
        Ppv2ac_in_ac = np.zeros_like(ppv)
        Ppv = np.empty_like(ppv)  # DC power output of the PV generator

        if pvmod:  # ppv: Normalized DC power output of the PV generator in kW/kWp
            Ppv = ppv * parameter['P_PV'] * 1000
        else:
            Ppv = ppv

        # DC power output of the PV generator taking into account the maximum
        # DC input power of the PV2AC conversion pathway
        Ppv = np.minimum(Ppv, parameter['P_PV2AC_in'] * 1000)

        # Residual power

        # Power demand on the AC side
        Pac = pl + parameter['P_PERI_AC']

        # Normalized AC output power of the PV2AC conversion pathway to cover the AC
        # power demand
        ppv2ac = np.minimum(
            Pac, parameter['P_PV2AC_out'] * 1000) / parameter['P_PV2AC_out'] / 1000

        # Target DC input power of the PV2AC conversion pathway
        Ppv2ac_in_ac = np.minimum(Pac, parameter['P_PV2AC_out'] * 1000) + (
            parameter['PV2AC_a_out'] * ppv2ac**2 + parameter['PV2AC_b_out'] * ppv2ac + parameter['PV2AC_c_out'])

        # Normalized DC input power of the PV2AC conversion pathway TODO 1
        ppv2ac = Ppv / parameter['P_PV2AC_in'] / 1000

        # Target AC output power of the PV2AC conversion pathway
        Ppv2ac_out = np.maximum(
            0, Ppv - (parameter['PV2AC_a_in'] * ppv2ac**2 + parameter['PV2AC_b_in'] * ppv2ac + parameter['PV2AC_c_in']))

        # Residual power for battery charging
        Prpv = Ppv - Ppv2ac_in_ac

        # Residual power for battery discharging
        Pr = Ppv2ac_out - Pac

        return Pr, Prpv, Ppv, ppv2ac, Ppv2ac_out

    # Maximize self consumption for PV-coupled systems
    elif parameter['Top'] == 'PV':
        # Preallocation
        # Pbat = np.zeros_like(ppv) # DC power of the battery in W
        # soc = np.zeros_like(ppv) # State of charge of the battery
        # Ppv2ac_out = np.zeros_like(ppv) # Output power of the PV2AC conversion pathway in W
        # Ppv2bat_in = np.zeros_like(ppv) # Input power of the PV2BAT conversion pathway in W
        # Pbat2pv_out = np.zeros_like(ppv) # Output power of the BAT2PV conversion pathway in W
        # Ppvbs = np.zeros_like(ppv) # AC power of the PV-battery system in W
        Ppv = np.empty_like(ppv)  # DC power output of the PV generator
        # Additional power consumption of other system components (e.g. AC power meter) in W
        Pperi = np.ones_like(ppv) * parameter['P_PERI_AC']
        # dt = 1 # Time increment in s
        # th = 0 # Start threshold for the recharging of the battery
        # soc0 = 0 # State of charge of the battery in the first time step

        # DC power output of the PV generator
        if pvmod:  # ppv: Normalized DC power output of the PV generator in kW/kWp
            Ppv = ppv * parameter['P_PV'] * 1000

        else:  # ppv: DC power output of the PV generator in W
            Ppv = ppv

        # Power demand on the AC side
        Pac = pl + Pperi

        return Pac, Ppv, Pperi


@nb.jit(nopython=True)
def BatMod_AC(d, _dt, _soc0, _soc, _Pr, _Pbs0, _Pbs, _Pbat):
    """Performance Simulation function for AC-coupled battery systems

    :param d: array containing parameters
    :type d: numpy array
    :param dt: time step width
    :type dt: integer
    :param soc0: state of charge in the previous time step
    :type soc0: float
    :param Pr: residual power
    :type Pr: numpy array
    :param Pbs0: AC-power of the battery system in the previous time step
    :type Pbs0: float
    :param Pbs: AC-power of the battery syste
    :type Pbs: numpy array
    :param Pbat: DC-power oof the battery
    :type Pbat: numpy array
    """
    # Loading of particular variables
    _E_BAT = d[0]
    _eta_BAT = d[1]
    _t_CONSTANT = d[2]
    _P_SYS_SOC0_DC = d[3]
    _P_SYS_SOC0_AC = d[4]
    _P_SYS_SOC1_DC = d[5]
    _P_SYS_SOC1_AC = d[6]
    _AC2BAT_a_in = d[7]
    _AC2BAT_b_in = d[8]
    _AC2BAT_c_in = d[9]
    _BAT2AC_a_out = d[10]
    _BAT2AC_b_out = d[11]
    _BAT2AC_c_out = d[12]
    _P_AC2BAT_DEV = d[13]
    _P_BAT2AC_DEV = d[14]
    _P_BAT2AC_out = d[15]
    _P_AC2BAT_in = d[16]
    _t_DEAD = int(round(d[17]))
    _SOC_h = d[18]

    _P_AC2BAT_min = _AC2BAT_c_in
    _P_BAT2AC_min = _BAT2AC_c_out

    # Correction factor to avoid over charge and discharge the battery
    corr = 0.1

    # Initialization of particular variables

    _tde = _t_CONSTANT > 0  # Binary variable to activate the first-order time delay element
    # Factor of the first-order time delay element
    _ftde = 1 - np.exp(-_dt / _t_CONSTANT)

    # First time step with regard to the dead time of the system control
    _tstart = np.maximum(2, 1 + _t_DEAD)
    _tend = int(_Pr.size)
    _th = 0

    # Capacity of the battery, conversion from kWh to Wh
    _E_BAT *= 1000

    # Effiency of the battery in percent
    _eta_BAT /= 100

    # Check if the dead or settling time can be ignored and set flags accordingly
    if _dt >= (3 * _t_CONSTANT) or _tend == 1:
        _tstart = 1
        T_DEAD = False
    else:
        T_DEAD = True

    if _dt >= _t_DEAD + 3 * _t_CONSTANT:
        SETTLING = False
    else:
        SETTLING = True

    for t in range(_tstart - 1, _tend):

        # Energy content of the battery in the previous time step
        E_b0 = _soc0 * _E_BAT

        # Calculate the AC power of the battery system from the residual power
        # with regard to the dead time of the system control
        if T_DEAD:
            P_bs = _Pr[t - _t_DEAD]
        else:
            P_bs = _Pr[t]

        # Check if the battery holds enough unused capacity for charging or discharging
        # Estimated amount of energy in Wh that is supplied to or discharged from the storage unit.
        E_bs_est = P_bs * _dt / 3600
        
        # Reduce P_bs to avoid over charging of the battery
        if E_bs_est > 0 and E_bs_est > (_E_BAT - E_b0):
            P_bs = (_E_BAT - E_b0) * 3600 / _dt
        # When discharging take the correction factor into account
        elif E_bs_est < 0 and np.abs(E_bs_est) > (E_b0):
            P_bs = (E_b0 * 3600 / _dt) * (1-corr)
        
        # Adjust the AC power of the battery system due to the stationary
        # deviations taking the minimum charging and discharging power into
        # account
        if P_bs > _P_AC2BAT_min:
            P_bs = np.maximum(_P_AC2BAT_min, P_bs + _P_AC2BAT_DEV)

        elif P_bs < -_P_BAT2AC_min:
            P_bs = np.minimum(-_P_BAT2AC_min, P_bs - _P_BAT2AC_DEV)

        else:
            P_bs = 0

        # Limit the AC power of the battery system to the rated power of the
        # battery converter
        P_bs = np.maximum(-_P_BAT2AC_out * 1000,
                          np.minimum(_P_AC2BAT_in * 1000, P_bs))

        # Adjust the AC power of the battery system due to the settling time
        # (modeled by a first-order time delay element) Hier hat der Schritt vorher eine Null?
        # Muss der vorherige Wert mit übergeben werden?
        if SETTLING:
            if t > 0:
                P_bs = _tde * _Pbs[t-1] + _tde * (P_bs - _Pbs[t-1]) * _ftde + P_bs * (not _tde)
            else:
                P_bs = _tde * _Pbs0 + _tde * (P_bs - _Pbs0) * _ftde + P_bs * (not _tde)

        # Decision if the battery should be charged or discharged
        if P_bs > 0 and _soc0 < 1 - _th * (1 - _SOC_h):
            # The last term th*(1-SOC_h) avoids the alternation between
            # charging and standby mode due to the DC power consumption of the
            # battery converter when the battery is fully charged. The battery
            # will not be recharged until the SOC falls below the SOC-threshold
            # (SOC_h) for recharging from PV.

            # Normalized AC power of the battery system
            p_bs = P_bs / _P_AC2BAT_in / 1000

            # DC power of the battery affected by the AC2BAT conversion losses
            # of the battery converter
            P_bat = np.maximum(
                0, P_bs - (_AC2BAT_a_in * p_bs * p_bs + _AC2BAT_b_in * p_bs + _AC2BAT_c_in))

        elif P_bs < 0 and _soc0 > 0:

            # Normalized AC power of the battery system
            p_bs = np.abs(P_bs / _P_BAT2AC_out / 1000)

            # DC power of the battery affected by the BAT2AC conversion losses
            # of the battery converter
            P_bat = P_bs - (_BAT2AC_a_out * p_bs * p_bs +
                            _BAT2AC_b_out * p_bs + _BAT2AC_c_out)

        else:  # Neither charging nor discharging of the battery

            # Set the DC power of the battery to zero
            P_bat = 0

        # Decision if the standby mode is active
        if P_bat == 0 and _soc0 <= 0:  # Standby mode in discharged state

            # DC and AC power consumption of the battery converter
            P_bat = -np.maximum(0, _P_SYS_SOC0_DC)
            P_bs = _P_SYS_SOC0_AC

        elif P_bat == 0 and _soc0 > 0:  # Standby mode in fully charged state

            # DC and AC power consumption of the battery converter
            P_bat = -np.maximum(0, _P_SYS_SOC1_DC)
            P_bs = _P_SYS_SOC1_AC

        # Transfer the realized AC power of the battery system and
        # the DC power of the battery
        _Pbs[t] = P_bs
        _Pbs0 = P_bs
        _Pbat[t] = P_bat

        # Change the energy content of the battery from Ws to Wh conversion
        if P_bat > 0:
            E_b = E_b0 + P_bat * np.sqrt(_eta_BAT) * _dt / 3600

        elif P_bat < 0:
            E_b = E_b0 + P_bat / np.sqrt(_eta_BAT) * _dt / 3600

        else:
            E_b = E_b0

        # Calculate the state of charge of the battery
        _soc0 = E_b / (_E_BAT)
        _soc[t] = _soc0

        # Adjust the hysteresis threshold to avoid alternation
        # between charging and standby mode due to the DC power
        # consumption of the battery converter.
        if _th and _soc[t] > _SOC_h or _soc[t] > 1:
            _th = True
        else:
            _th = False

    return _Pbat, _Pbs, _soc, _soc0, _Pbs0

@nb.jit(nopython=True)
def BatMod_DC(d, _dt, _soc0, _soc, _Pr, _Prpv,  _Ppv, _Ppv2bat_in0, _Ppv2bat_in, _Pbat2ac_out0, _Pbat2ac_out, _Ppv2ac_out0, _Ppv2ac_out, _Ppvbs, _Pbat):
    """Performance simulation function for DC-coupled battery systems

    :param d: array containing parameters
    :type d: numpy array
    :param dt: time step width
    :type dt: integer
    :param soc0: state of charge in the previous time step
    :type soc0: float
    :param Pr: residual power
    :type Pr: numpy array
    :param Prpv: residual power of the PV-system
    :type Prpv: numpy array
    :param Ppv: PV-power
    :type Ppv: numpy array
    :param Ppv2bat_in0: AC input power of the battery system in the previous time step
    :type Ppv2bat_in0: float
    :param Ppv2bat_in: AC input power of the battery system
    :type Ppv2bat_in: numpy array
    :param Pbat2ac_out0: AC output power of the battery system in the previous time step
    :type Pbat2ac_out0: float
    :param Pbat2ac_out: AC output power of the battery system
    :type Pbat2ac_out: numpy array
    :param Ppv2ac_out0: AC output power of the PV inverter in the previous time step
    :type Ppv2ac_out0: float
    :param Ppv2ac_out: AC output power of the PV inverter
    :type Ppv2ac_out: numpy array
    :param Ppvbs: AC power from the PV system to the battery system
    :type Ppvbs: numpy array
    :param Pbat: DC power of the battery
    :type Pbat: float
    """

    _E_BAT = d[0]
    _P_PV2AC_in = d[1]
    _P_PV2AC_out = d[2]
    _P_PV2BAT_in = d[3]
    _P_BAT2AC_out = d[4]
    _PV2AC_a_in = d[5]
    _PV2AC_b_in = d[6]
    _PV2AC_c_in = d[7]
    _PV2BAT_a_in = d[8]
    _PV2BAT_b_in = d[9]
    _BAT2AC_a_out = d[10]
    _BAT2AC_b_out = d[11]
    _BAT2AC_c_out = d[12]
    _eta_BAT = d[13]
    _SOC_h = d[14]
    _P_PV2BAT_DEV = d[15]
    _P_BAT2AC_DEV = d[16]
    _t_DEAD = int(round(d[17]))
    _t_CONSTANT = d[18]
    _P_SYS_SOC1_DC = d[19]
    _P_SYS_SOC0_AC = d[20]
    _P_SYS_SOC0_DC = d[21]
    _P_PV2AC_min = _PV2AC_c_in

    # Capacity of the battery, conversion from kWh to Wh
    _E_BAT *= 1000

    # Effiency of the battery in percent
    _eta_BAT /= 100

    # Initialization of particular variables
    # _P_PV2AC_min = _parameter['PV2AC_c_in'] # Minimum input power of the PV2AC conversion pathway
    _tde = _t_CONSTANT > 0  # Binary variable to activate the first-order time delay element
    # Factor of the first-order time delay element
    _ftde = 1 - np.exp(-_dt / _t_CONSTANT)
    # First time step with regard to the dead time of the system control
    _tstart = np.maximum(2, 1 + _t_DEAD)
    _tend = int(_Pr.size)
    _th = 0
    corr = 0.1

    # Check if the dead or settling time can be ignored and set flags accordingly
    if _dt >= (3 * _t_CONSTANT) or _tend == 1:
        _tstart = 1
        T_DEAD = False
    else:
        T_DEAD = True

    if _dt >= _t_DEAD + 3 * _t_CONSTANT:
        SETTLING = False
    else:
        SETTLING = True

    for t in range(_tstart - 1, _tend):
        # Energy content of the battery in the previous time step
        E_b0 = _soc0 * _E_BAT

        # Residual power with regard to the dead time of the system control

        if T_DEAD:
            P_rpv = _Prpv[t - _t_DEAD]
            P_r = _Pr[t - _t_DEAD]

        else:
            P_rpv = _Prpv[t]
            P_r = _Pr[t]

        # Check if the battery holds enough unused capacity for charging or discharging
        # Estimated amount of energy that is supplied to or discharged from the storage unit.
        E_bs_rpv = P_rpv * _dt / 3600
        E_bs_r = P_r * _dt / 3600
        
        # Reduce P_bs to avoid over charging of the battery 
        if E_bs_rpv > 0 and E_bs_rpv > (_E_BAT - E_b0):
            P_rpv = (_E_BAT - E_b0) * 3600 / _dt
        # When discharging take the correction factor into account
        elif E_bs_r < 0 and np.abs(E_bs_r) > (E_b0):
            P_r = ((E_b0) * 3600 / _dt) * (1-corr)
        
        # Decision if the battery should be charged or discharged
        if P_rpv > 0 and _soc0 < 1 - _th * (1 - _SOC_h):
            '''
            The last term th*(1-SOC_h) avoids the alternation between
            charging and standby mode due to the DC power consumption of the
            battery converter when the battery is fully charged. The battery
            will not be recharged until the SOC falls below the SOC-threshold
            (SOC_h) for recharging from PV.
            '''
            # Charging power
            P_pv2bat_in = P_rpv

            # Adjust the charging power due to the stationary deviations
            P_pv2bat_in = np.maximum(0, P_pv2bat_in + _P_PV2BAT_DEV)

            # Limit the charging power to the maximum charging power
            P_pv2bat_in = np.minimum(P_pv2bat_in, _P_PV2BAT_in * 1000)

            # Adjust the charging power due to the settling time
            # (modeled by a first-order time delay element)
            if SETTLING:
                if t > 0:
                    P_pv2bat_in = _tde * _Ppv2bat_in[(t-1)] + _tde * (
                        P_pv2bat_in - _Ppv2bat_in[(t-1)]) * _ftde + P_pv2bat_in * (not _tde)
                else:
                    P_pv2bat_in = _tde * _Ppv2bat_in0 + _tde * \
                        (P_pv2bat_in - _Ppv2bat_in0) * \
                        _ftde + P_pv2bat_in * (not _tde)
                        
            # Limit the charging power to the current power output of the PV generator
            P_pv2bat_in = np.minimum(P_pv2bat_in, _Ppv[t])

            # Normalized charging power
            ppv2bat = P_pv2bat_in / _P_PV2BAT_in / 1000

            # DC power of the battery affected by the PV2BAT conversion losses
            # (the idle losses of the PV2BAT conversion pathway are not taken
            # into account)
            P_bat = np.maximum(
                0, P_pv2bat_in - (_PV2BAT_a_in * ppv2bat**2 + _PV2BAT_b_in * ppv2bat))

            # Realized DC input power of the PV2AC conversion pathway
            P_pv2ac_in = _Ppv[t] - P_pv2bat_in

            # Normalized DC input power of the PV2AC conversion pathway
            _ppv2ac = P_pv2ac_in / _P_PV2AC_in / 1000

            # Realized AC power of the PV-battery system
            P_pv2ac_out = np.maximum(
                0, P_pv2ac_in - (_PV2AC_a_in * _ppv2ac**2 + _PV2AC_b_in * _ppv2ac + _PV2AC_c_in))
            P_pvbs = P_pv2ac_out

            # Transfer the final values
            _Ppv2ac_out[t] = P_pv2ac_out
            _Ppv2bat_in0 = P_pv2bat_in
            _Ppv2bat_in[t] = P_pv2bat_in

        elif P_rpv < 0 and _soc0 > 0:

            # Discharging power
            P_bat2ac_out = P_r * -1

            # Adjust the discharging power due to the stationary deviations
            P_bat2ac_out = np.maximum(0, P_bat2ac_out + _P_BAT2AC_DEV)

            # Adjust the discharging power to the maximum discharging power
            P_bat2ac_out = np.minimum(P_bat2ac_out, _P_BAT2AC_out * 1000)

            # Adjust the discharging power due to the settling time
            # (modeled by a first-order time delay element)
            if SETTLING:
                if t > 0:
                    P_bat2ac_out = _tde * _Pbat2ac_out[t-1] + _tde * (
                        P_bat2ac_out - _Pbat2ac_out[t-1]) * _ftde + P_bat2ac_out * (not _tde)
                else:
                    P_bat2ac_out = _tde * _Pbat2ac_out0 + _tde * \
                        (P_bat2ac_out - _Pbat2ac_out0) * \
                        _ftde + P_bat2ac_out * (not _tde)

            # Limit the discharging power to the maximum AC power output of the PV-battery system
            P_bat2ac_out = np.minimum(
                _P_PV2AC_out * 1000 - _Ppv2ac_out[t], P_bat2ac_out)

            # Normalized discharging power
            ppv2bat = P_bat2ac_out / _P_BAT2AC_out / 1000

            # DC power of the battery affected by the BAT2AC conversion losses
            # (if the idle losses of the PV2AC conversion pathway are covered by
            # the PV generator, the idle losses of the BAT2AC conversion pathway
            # are not taken into account)
            if _Ppv[t] > _P_PV2AC_min:
                P_bat = -1 * (P_bat2ac_out + (_BAT2AC_a_out *
                                              ppv2bat**2 + _BAT2AC_b_out * ppv2bat))
            else:
                P_bat = -1 * (P_bat2ac_out + (_BAT2AC_a_out * ppv2bat **
                                              2 + _BAT2AC_b_out * ppv2bat + _BAT2AC_c_out)) + _Ppv[t]

            # Realized AC power of the PV-battery system
            P_pvbs = _Ppv2ac_out[t] + P_bat2ac_out

            # Transfer the final values
            _Pbat2ac_out0 = P_bat2ac_out
            _Pbat2ac_out[t] = P_bat2ac_out

        else:  # Neither charging nor discharging of the battery

            # Set the DC power of the battery to zero
            P_bat = 0

            # Realized AC power of the PV-battery system
            P_pvbs = _Ppv2ac_out[t]

        # Decision if the standby mode is active
        if P_bat == 0 and P_pvbs == 0 and _soc0 <= 0:  # Standby mode in discharged state

            # DC and AC power consumption of the PV-battery inverter
            P_bat = -np.maximum(0, _P_SYS_SOC0_DC)
            P_pvbs = -_P_SYS_SOC0_AC

        elif P_bat == 0 and P_pvbs > 0 and _soc0 > 0:  # Standby mode in fully charged state

            # DC power consumption of the PV-battery inverter
            P_bat = -np.maximum(0, _P_SYS_SOC1_DC)

        # Transfer the realized AC power of the PV-battery system and the DC power of the battery
        _Ppvbs[t] = P_pvbs
        _Pbat[t] = P_bat

        # Change the energy content of the battery Wx to Wh conversion
        if P_bat > 0:
            E_b = E_b0 + P_bat * np.sqrt(_eta_BAT) * _dt / 3600
        elif P_bat < 0:
            E_b = E_b0 + P_bat / np.sqrt(_eta_BAT) * _dt / 3600
        else:
            E_b = E_b0

        # Calculate the state of charge of the battery
        _soc0 = E_b / _E_BAT
        _soc[t] = _soc0

        # Adjust the hysteresis threshold to avoid alternation between charging
        # and standby mode due to the DC power consumption of the
        # PV-battery inverter
        if _th and _soc[t] > _SOC_h or _soc[t] > 1:
            _th = True
        else:
            _th = False

    return _Ppv2ac_out, _Ppv2bat_in, _Ppv2bat_in0, _Pbat2ac_out, _Pbat2ac_out0, _Ppvbs, _Pbat, _soc, _soc0

@nb.jit(nopython=True)
def BatMod_PV(d, _dt, _soc0, _soc, _Ppv, _Pac, _Ppv2bat_in0, _Ppv2bat_in, _Ppv2ac_out, _Pbat2pv_out0, _Pbat2pv_out, _Ppvbs, _Pbat):
    """Performance simulation function for PV-coupled battery systems

    :param d: array containing parameters
    :type d: numpy array

    :param dt: time step width
    :type dt: integer

    :param soc0: state of charge of the battery in the previous time step
    :type soc0: float

    :param soc: state of charge of the battery
    :type soc: numpy array

    :param Pr: residual power
    :type Pr: numpy array

    :param Ppv: PV-power
    :type Ppv: numpy array

    :param Pac: AC output power of the PV inverter
    :type Pac: numpy array

    :param Ppv2bat_in: AC input power of the battery system
    :type Ppv2bat_in: numpy array

    :param Ppv2bat_in0: AC input power of the battery system in the previous time step
    :type Ppv2bat_in0: float

    :param Pbat2pv_out0: AC output power of the battery system in the previous time step
    :type Pbat2pv_out0: float

    :param Pbat2pv_out: AC output power of the battery system
    :type Pbat2pv_out: numpy array

    :param Ppvbs: AC power from the PV system to the battery system
    :type Ppvbs: numpy array

    :param Pbat: DC power of the battery
    :type Pbat: float
    """

    # Initialization of particular variables

    _E_BAT = d[0]
    _P_PV2AC_in = d[1]
    _P_PV2AC_out = d[2]
    _P_PV2BAT_in = d[3]
    _P_BAT2PV_out = d[4]
    _PV2AC_a_in = d[5]
    _PV2AC_b_in = d[6]
    _PV2AC_c_in = d[7]
    _PV2BAT_a_in = d[8]
    _PV2BAT_b_in = d[9]
    _PV2BAT_c_in = d[10]
    _PV2AC_a_out = d[11]
    _PV2AC_b_out = d[12]
    _PV2AC_c_out = d[13]
    _BAT2PV_a_out = d[14]
    _BAT2PV_b_out = d[15]
    _BAT2PV_c_out = d[16]
    _eta_BAT = d[17]
    _SOC_h = d[18]
    _P_PV2BAT_DEV = d[19]
    _P_BAT2AC_DEV = d[20]
    _P_SYS_SOC1_DC = d[21]
    _P_SYS_SOC0_AC = d[22]
    _P_SYS_SOC0_DC = d[23]
    _t_DEAD = int(round(d[24]))
    _t_CONSTANT = d[25]

    # Correction factor to avoid over charge and discharge the battery
    corr = 0.1

    _P_PV2BAT_min = _PV2BAT_c_in  # Minimum DC charging power
    _P_BAT2PV_min = _BAT2PV_c_out  # Minimum DC discharging power

    # Initialization of particular variables
    _tde = _t_CONSTANT > 0  # Binary variable to activate the first-order time delay element
    # Factor of the first-order time delay element
    _ftde = 1 - np.exp(-_dt / _t_CONSTANT)
    # First time step with regard to the dead time of the system control
    _tstart = np.maximum(2, 1 + _t_DEAD)
    _tend = int(_Ppv.size)
    _th = 0

    _E_BAT *= 1000  # Conversion from W to kW

    _eta_BAT /= 100

    # Check if the dead or settling time can be ignored and set flags accordingly
    if _dt >= (3 * _t_CONSTANT) or _tend == 1:
        _tstart = 1
        T_DEAD = False
    else:
        T_DEAD = True

    if _dt >= _t_DEAD + 3 * _t_CONSTANT:
        SETTLING = False
    else:
        SETTLING = True

    for t in range(_tstart - 1, _tend):

        # Energy content of the battery in the previous time step
        E_b0 = _soc0 * _E_BAT

        # Target AC output power of the PV-battery system to cover the AC power demand
        if T_DEAD:
            P_pvbs = np.minimum(_Pac[t - _t_DEAD], _P_PV2AC_out * 1000)
        else:
            P_pvbs = np.minimum(_Pac[t], _P_PV2AC_out * 1000)

        # Normalized AC output power of the PV2AC conversion pathway
        ppv2ac = P_pvbs / _P_PV2AC_out / 1000

        # Target DC input power of the PV2AC conversion pathway
        P_pv2ac_in = P_pvbs + (_PV2AC_a_out * ppv2ac **
                               2 + _PV2AC_b_out * ppv2ac + _PV2AC_c_out)

        # Residual power
        if T_DEAD:
            P_rpv = _Ppv[t - _t_DEAD] - P_pv2ac_in
        else:
            P_rpv = _Ppv[t] - P_pv2ac_in

        # Check if the battery holds enough unused capacity for charging or discharging
        # Estimated amount of energy that is supplied to or discharged from the storage unit.
        E_bs_rpv = P_rpv * _dt / 3600
        
        # Reduce P_bs to avoid over charging of the battery
        if E_bs_rpv > 0 and E_bs_rpv > (_E_BAT - E_b0):
            P_rpv = ((_E_BAT - E_b0) * 3600) / _dt
        # When charging take the correction factor into account
        elif E_bs_rpv < 0 and np.abs(E_bs_rpv) > (E_b0):
            P_rpv = ((E_b0) * 3600 / _dt) * (1-corr)
        
        # Decision if the battery should be charged or discharged
        if P_rpv > _P_PV2BAT_min and _soc0 < 1 - _th * (1 - _SOC_h):
            '''
            The last term th*(1-SOC_h) avoids the alternation between
            charging and standby mode due to the DC power consumption of the
            battery converter when the battery is fully charged. The battery
            will not be recharged until the SOC falls below the SOC-threshold
            (SOC_h) for recharging from PV.
            '''
            # Charging power
            P_pv2bat_in = P_rpv

            # Adjust the charging power due to stationary deviations
            P_pv2bat_in = np.maximum(0, P_pv2bat_in + _P_PV2BAT_DEV)

            # Limit the charging power to the maximum charging power
            P_pv2bat_in = np.minimum(P_pv2bat_in, _P_PV2BAT_in * 1000)

            # Adjust the charging power due to the settling time
            # (modeled by a first-order time delay element)
            if SETTLING:
                if t > 0:
                    P_pv2bat_in = _tde * _Ppv2bat_in[t-1] + _tde * (
                        P_pv2bat_in - _Ppv2bat_in[t-1]) * _ftde + P_pv2bat_in * (not _tde)
                else:
                    P_pv2bat_in = _tde * _Ppv2bat_in0 + _tde * \
                        (P_pv2bat_in - _Ppv2bat_in0) * \
                        _ftde + P_pv2bat_in * (not _tde)

            # Limit the charging power to the current power output of the PV generator
            P_pv2bat_in = np.minimum(P_pv2bat_in, _Ppv[t])

            # Normalized charging power
            ppv2bat = P_pv2bat_in / _P_PV2BAT_in / 1000

            # DC power of the battery
            P_bat = np.maximum(0, P_pv2bat_in - (_PV2BAT_a_in *
                                                 ppv2bat**2 + _PV2BAT_b_in * ppv2bat + _PV2BAT_c_in))

            # Realized DC input power of the PV2AC conversion pathway
            P_pv2ac_in = _Ppv[t] - P_pv2bat_in

            # Limit the DC input power of the PV2AC conversion pathway
            P_pv2ac_in = np.minimum(P_pv2ac_in, _P_PV2AC_in * 1000)

            # Recalculate Ppv(t) with limited PV2AC input power
            _Ppv[t] = P_pv2ac_in + P_pv2bat_in

            # Normalized DC input power of the PV2AC conversion pathway
            ppv2ac = P_pv2ac_in / _P_PV2AC_in / 1000

            # Realized AC power of the PV-battery system
            P_pv2ac_out = np.maximum(
                0, P_pv2ac_in - (_PV2AC_a_in * ppv2ac**2 + _PV2AC_b_in * ppv2ac + _PV2AC_c_in))
            P_pvbs = P_pv2ac_out

            # Transfer the final values
            _Ppv2ac_out[t] = P_pv2ac_out
            _Ppv2bat_in[t] = P_pv2bat_in

        elif P_rpv < -_P_BAT2PV_min and _soc0 > 0:
            # Target discharging power of the battery
            P_bat2pv_out = np.abs(P_rpv)

            # Adjust the discharging power due to the stationary deviations
            P_bat2pv_out = np.maximum(0, P_bat2pv_out + _P_BAT2AC_DEV)

            # Adjust the discharging power to the maximum discharging power
            P_bat2pv_out = np.minimum(P_bat2pv_out, _P_BAT2PV_out * 1000)

            # Adjust the discharging power due to the settling time
            # (modeled by a first-order time delay element)
            if SETTLING:
                if t > 0:
                    P_bat2pv_out = _tde * _Pbat2pv_out[t-1] + _tde * (P_bat2pv_out - _Pbat2pv_out[t-1]) * _ftde + P_bat2pv_out * (not _tde)
                else:
                    P_bat2pv_out = _tde * _Pbat2pv_out0 + _tde * (P_bat2pv_out - _Pbat2pv_out0) * _ftde + P_bat2pv_out * (not _tde)
            
            # Recalculate Ppv(t) with limited PV2AC input power
            _Ppv[t] = np.minimum(_P_PV2AC_in * 1000, _Ppv[t])

            # Limit the discharging power to the maximum AC power output of the PV-battery system
            P_bat2pv_out = np.minimum(_P_PV2AC_in * 1000 - _Ppv[t], P_bat2pv_out)

            # Normalized discharging power
            pbat2pv = P_bat2pv_out / _P_BAT2PV_out / 1000

            # DC power of the battery affected by the BAT2PV conversion losses
            P_bat = -1*(P_bat2pv_out+(_BAT2PV_a_out * pbat2pv**2 + _BAT2PV_b_out * pbat2pv + _BAT2PV_c_out))

            # Realized DC input power of the PV2AC conversion pathway
            P_pv2ac_in = _Ppv[t] + P_bat2pv_out

            # Normalized DC input power of the PV2AC conversion pathway
            ppv2ac = P_pv2ac_in / _P_PV2AC_in / 1000

            # AC power of the PV-battery system
            P_pvbs = np.maximum(0, P_pv2ac_in-(_PV2AC_a_in * ppv2ac**2 + _PV2AC_b_in * ppv2ac + _PV2AC_c_in))
            P_pv2ac_out = P_pvbs

            # Transfer the final values
            _Ppv2ac_out[t] = P_pv2ac_out
            _Pbat2pv_out[t] = P_bat2pv_out

        else:  # Neither charging nor discharging of the battery

            # Set the DC power of the battery to zero
            P_bat = 0

            # Limit the power output of the PV generator to the maximum input power
            # of the PV inverter
            _Ppv[t] = np.minimum(_Ppv[t], _P_PV2AC_in * 1000)

            # Normalized DC input power of the PV2AC conversion pathway
            ppv2ac = _Ppv[t] / _P_PV2AC_in / 1000

            # Realized AC power of the PV-battery system
            P_pvbs = np.maximum(0, _Ppv[t] - (_PV2AC_a_in * ppv2ac**2 + _PV2AC_b_in * ppv2ac + _PV2AC_c_in))

            # Transfer the final values
            _Ppv2ac_out[t] = P_pvbs

        # Decision if the standby mode is active
        if P_bat == 0 and _soc0 <= 0:  # Standby mode in discharged state

            # DC power consumption of the battery converter
            P_bat = -np.maximum(0, _P_SYS_SOC0_DC)
            if P_pvbs == 0:
                P_pvbs = -_P_SYS_SOC0_AC

        elif P_bat == 0 and P_pvbs > 0 and _soc0 > 0:  # Standby mode in fully charged state

            # DC power consumption of the battery converter
            P_bat = -np.maximum(0, _P_SYS_SOC1_DC)

        # Transfer the realized AC power of the battery system and
        # the DC power of the battery
        _Ppvbs[t] = P_pvbs
        _Pbat[t] = P_bat

        # Change the energy content of the battery Wx to Wh conversio
        if P_bat > 0:
            E_b = E_b0 + P_bat * np.sqrt(_eta_BAT) * _dt / 3600

        elif P_bat < 0:
            E_b = E_b0 + P_bat / np.sqrt(_eta_BAT) * _dt / 3600

        else:
            E_b = E_b0

        # Calculate the state of charge of the battery
        _soc0 = E_b / (_E_BAT)
        _soc[t] = _soc0

        # Adjust the hysteresis threshold to avoid alternation
        # between charging and standby mode due to the DC power
        # consumption of the battery converter.
        if _th and _soc[t] > _SOC_h or _soc[t] > 1:
            _th = True
        else:
            _th = False

    return _soc, _soc0, _Ppv, _Ppvbs, _Pbat, _Ppv2ac_out, _Pbat2pv_out, _Ppv2bat_in

def bat_res_mod(_parameter, _Pl, _Ppv, _Pbat, _dt, *args):
    """Function for calculating energy sums

    :param _parameter: parameter of the system
    :type _parameter: dict
    :param _Pl: load power
    :type _Pl: numpy array
    :param _Ppv: output power of the PV generator
    :type _Ppv: numpy array
    :param _Pbat: DC power of the battery
    :type _Pbat: numpy array
    :param _dt: time step width
    :type _dt: integer
    :return: energy sums
    :rtype: dict
    """
    _E = dict()

    if _parameter['Top'] == 'AC':  # AC-coupled systems

        _Ppvs = args[0]  # AC output power of the PV system
        _Pbs = args[1]  # AC power of the battery system
        # Additional power consumption of the other system components
        _Pperi = args[2]

    elif _parameter['Top'] == 'DC' or _parameter['Top'] == 'PV':  # DC- and PV-coupled systems

        _Ppv2ac = args[0]  # AC output power of the PV2AC conversion pathway
        _Ppv2bat_in = args[1]  # Input power of the PV2BAT conversion pathway
        _Ppvbs = args[2]  # AC power of the PV-battery system
        # Additional power consumption of the other system components
        _Pperi = args[3]

        _Ppv2ac_in = _Ppv - _Ppv2bat_in  # Input power of the PV2AC conversion pathway

    # Total load including the power consumption of the other system components
    _Plt = _Pl + _Pperi
    # DC input power of the battery (charged)
    _Pbatin = np.maximum(0, _Pbat)
    # DC output power of the battery (discharged)
    _Pbatout = np.minimum(0, _Pbat)
    # Maximum PV feed-in power
    _P_ac2g_max = _parameter['p_ac2g_max'] * _parameter['P_PV'] * 1000

    if _parameter['Top'] == 'AC':  # AC-coupled systems

        # Residual power without curtailment
        _Pr = _Ppvs - _Plt
        # AC input power of the battery system
        _Pac2bs = np.maximum(0, _Pbs)
        # AC output power of the battery system
        _Pbs2ac = np.minimum(0, _Pbs)
        # Negative residual power (residual load demand)
        _Prn = np.minimum(0, _Pr)
        # Positive residual power (surplus PV power)
        _Prp = np.maximum(0, _Pr)
        # Direct use of PV power by the load
        _Ppvs2l = np.minimum(_Ppvs, _Plt)
        # PV charging power
        _Ppvs2bs = np.minimum(_Prp, _Pac2bs)
        # Grid charging power
        _Pg2bs = np.maximum(_Pac2bs - _Prp, 0)
        # Grid supply power of the load
        _Pg2l = np.minimum(_Prn - _Pbs2ac, 0)
        # Battery supply power of the load
        _Pbs2l = np.maximum(_Prn, _Pbs2ac)
        # Battery feed-in power
        _Pbs2g = np.minimum(_Pbs2ac - _Prn, 0)
        # PV feed-in power including curtailment
        _Ppvs2g = np.minimum(np.maximum(_Prp - _Pac2bs, 0), _P_ac2g_max)
        # Power demand from the grid
        _Pg2ac = _Pg2l - _Pg2bs
        # Feed-in power to the grid
        _Pac2g = _Ppvs2g - _Pbs2g
        # Grid power
        _Pg = _Pac2g + _Pg2ac
        # Curtailed PV power (AC output power)
        _Pct = np.maximum(_Prp - _Pac2bs, 0) - _Ppvs2g
        # AC output power of the PV system including curtailment
        _Ppvs = _Ppvs - _Pct
        # Residual power including curtailment
        _Pr = _Ppvs - _Plt
        # Index for PV curtailment
        _idx = np.where(_Pct > 0)[0]

        for i in range(len(_idx)):

            _tct = _idx[i]
            # Normalized output power of the PV inverter
            _ppvinvout = _Ppvs[_tct] / _parameter['P_PV2AC_out'] / 1000
            # DC output power of the PV generator taking into account the
            # conversion and curtailment losses
            _Ppv[_tct] = _Ppvs[_tct] + (_parameter['PV2AC_a_out'] * _ppvinvout **
                                        2 + _parameter['PV2AC_b_out'] * _ppvinvout + _parameter['PV2AC_c_out'])

    elif _parameter['Top'] == 'DC' or _parameter['Top'] == 'PV':  # DC- and PV-coupled systems

        # Grid power demand of the PV-battery system
        _Pg2pvbs = np.minimum(0, _Ppvbs)
        # AC input power of the PV-battery system
        _Pac2pvbs = _Pg2pvbs
        # AC output power of the PV-battery system
        _Ppvbs2ac = np.maximum(0, _Ppvbs)
        # Load supply power by the PV-battery system
        _Ppvbs2l = np.minimum(_Plt, _Ppvbs2ac)
        # Load supply power by the grid
        _Pg2l = _Plt - _Ppvbs2l
        # Direct use of PV power by the load
        _Ppv2l = np.minimum(_Plt, _Ppv2ac)
        # PV feed-in power including curtailment
        _Ppv2g = np.minimum(_Ppv2ac - _Ppv2l, _P_ac2g_max)
        # Curtailed PV power (AC output power)
        _Pct = _Ppv2ac - _Ppv2l - _Ppv2g

        if np.sum(_Pct) > 0:
            # Power of the PV-battery system including curtailment
            _Ppvbs = _Ppvbs - _Pct
            # AC output power of the PV-battery system including curtailment
            _Ppvbs2ac = np.maximum(0, _Ppvbs)
            # AC output power of the PV2AC conversion pathway including curtailment
            _Ppv2ac = _Ppv2ac - _Pct
            # Index for PV curtailment
            _idx = np.where(_Pct > 0)[0]

            for i in range(len(_idx)):

                _tct = _idx[i]
                # Specific AC output power of the PV2AC conversion pathway
                _ppv2ac = _Ppv2ac[_tct] / _parameter['P_PV2AC_out'] / 1000
                # DC input power of the PV2AC conversion pathway including curtailment
                _Ppv2ac_in[_tct] = _Ppv2ac[_tct] + (_parameter['PV2AC_a_out'] * _ppv2ac **
                                                    2 + _parameter['PV2AC_b_out'] * _ppv2ac + _parameter['PV2AC_c_out'])

            # DC output power of the PV generator including curtailment
            _Ppv = _Ppv2ac_in + _Ppv2bat_in

        # Grid power including curtailment
        _Pg = _Ppvbs-_Plt
        # Feed-in power to the grid including curtailment
        _Pac2g = np.maximum(0, _Pg)
        # Power demand from the grid
        _Pg2ac = np.minimum(0, _Pg)

    # Energy sums in MWH

    # Electrical demand including the energy consumption of the other system components
    _E['El'] = np.sum(np.abs(_Plt)) * _dt / 3.6e9
    # DC output of the PV generator including curtailment
    _E['Epv'] = np.sum(np.abs(_Ppv)) * _dt / 3.6e9
    # DC input of the battery (charged)
    _E['Ebatin'] = np.sum(np.abs(_Pbatin)) * _dt / 3.6e9
    # DC output of the battery (discharged)
    _E['Ebatout'] = np.sum(np.abs(_Pbatout)) * _dt / 3.6e9
    # Grid feed-in
    _E['Eac2g'] = np.sum(np.abs(_Pac2g)) * _dt / 3.6e9
    # Grid demand
    _E['Eg2ac'] = np.sum(np.abs(_Pg2ac)) * _dt / 3.6e9
    # Load supply by the grid
    _E['Eg2l'] = np.sum(np.abs(_Pg2l)) * _dt / 3.6e9
    # Demand of the other system components
    _E['Eperi'] = np.sum(np.abs(_Pperi)) * _dt / 3.6e9
    # Curtailed PV energy
    _E['Ect'] = np.sum(np.abs(_Pct)) * _dt / 3.6e9

    if _parameter['Top'] == 'AC':  # AC-coupled systems

        # AC output of the PV system including curtailment
        _E['Epvs'] = np.sum(np.abs(_Ppvs)) * _dt / 3.6e9
        # AC input of the battery system
        _E['Eac2bs'] = np.sum(np.abs(_Pac2bs)) * _dt / 3.6e9
        # AC output of the battery system
        _E['Ebs2ac'] = np.sum(np.abs(_Pbs2ac)) * _dt / 3.6e9
        # Direct use of PV energy
        _E['Epvs2l'] = np.sum(np.abs(_Ppvs2l)) * _dt / 3.6e9
        # PV charging
        _E['Epvs2bs'] = np.sum(np.abs(_Ppvs2bs)) * _dt / 3.6e9
        # Grid charging
        _E['Eg2bs'] = np.sum(np.abs(_Pg2bs)) * _dt / 3.6e9
        # PV feed-in
        _E['Epvs2g'] = np.sum(np.abs(_Ppvs2g)) * _dt / 3.6e9
        # Load supply by the battery system
        _E['Ebs2l'] = np.sum(np.abs(_Pbs2l)) * _dt / 3.6e9
        # Battery feed-in
        _E['Ebs2g'] = np.sum(np.abs(_Pbs2g)) * _dt / 3.6e9

    elif _parameter['Top'] == 'DC' or _parameter['Top'] == 'PV':  # DC- and PV-coupled systems

        # Grid demand of the PV-battery system
        _E['Eg2pvbs'] = np.sum(np.abs(_Pg2pvbs)) * _dt / 3.6e9
        # AC input of the PV-battery system
        _E['Eac2pvbs'] = np.sum(np.abs(_Pac2pvbs)) * _dt / 3.6e9
        # AC output of the PV-battery system
        _E['Epvbs2ac'] = np.sum(np.abs(_Ppvbs2ac)) * _dt / 3.6e9
        # Load supply by the PV-battery system
        _E['Epvbs2l'] = np.sum(np.abs(_Ppvbs2l)) * _dt / 3.6e9

    return _E

def load_parameter(fname, col_name):
    """Loads system parameter from excel file

    :param fname: Path to the excel file
    :type fname: string
    :param col_name: Column to read data from
    :type col_name: string
    :return: Dictionary holding parameters from the Excel sheet
    :rtype: dict
    """
    _version = '0.1'
    
    wb = load_workbook(fname, data_only=True)
    ws = wb['Data'] # Load Data sheet of excel file

    # read keys and values from Excel sheet
    keys = (c.value for c in ws['E'][1:])
    values = (c.value if c.value != 'ns' else None for c in ws[col_name][1:])
    
    parameter = dict(zip(keys, values))

    # deletes entries where key is None
    del parameter[None]

    # Assign specific parameters
    parameter['P_PV2AC_out_PVINV'] = ws[col_name][15].value
    parameter['P_PV2AC_out'] = ws[col_name][24].value
    parameter['P_AC2BAT_in_DCC'] = ws[col_name][25].value
    parameter['P_AC2BAT_in'] = ws[col_name][26].value
    parameter['P_BAT2AC_out'] = ws[col_name][27].value
    parameter['P_BAT2AC_out_DCC'] = ws[col_name][28].value

    # Set refrence case values to boolean
    if parameter['ref_1'] == 'yes':
        parameter['ref_1'] = True
    elif parameter['ref_1'] == 'no':
        parameter['ref_1'] = False
    
    if parameter['ref_2'] == 'yes':
        parameter['ref_2'] = True
    elif parameter['ref_2'] == 'no':
        parameter['ref_2'] = False

    # Specific parameters of DC-coupled systems
    if parameter['Top'] == 'DC':
        parameter['P_AC2BAT_in'] = parameter['P_AC2BAT_in_DCC'] # Nominal charging power (AC) in kW
        parameter['P_BAT2AC_out'] = parameter['P_BAT2AC_out_DCC']
    
    # Specific parameters of PV inverters and AC-coupled systems
    if parameter['Top'] == 'PVINV' or parameter['Top'] == 'AC' and parameter['P_PV2AC_out_PVINV'] is not None:
        parameter['P_PV2AC_out'] = parameter['P_PV2AC_out_PVINV']
    
    # Specific parameters of PV-coupled systems
    if parameter['Top'] == 'PV':
        parameter['P_BAT2PV_in'] = parameter['P_BAT2AC_in']
        parameter['P_BAT2AC_out'] = parameter['P_BAT2AC_out_DCC']

    # replace 'ns', 'o' and 'c' entries to None
    for key, value in parameter.items():
        if value == 'ns' or value == 'o' or value == 'c' or value == ' ':
            parameter[key] = None

    # Convert to kW
    convert_to_kw = ['P_PV2AC_in', 'P_PV2AC_out_PVINV','P_PV2AC_out','P_AC2BAT_in_DCC','P_AC2BAT_in','P_BAT2AC_out',
             'P_BAT2AC_out_DCC','P_PV2BAT_in','P_BAT2PV_out','P_PV2BAT_out','P_BAT2AC_in']

    for par in convert_to_kw:
        if parameter[par]:
            parameter[par] /= 1000
        
    return parameter

def eta2abc(parameter):
    """Function to calculate the parameters of the power loss functions (quadratic equations) from the path efficiencies

    :param parameter: Holds parameters of the system
    :type parameter: dict
    :return: Dictionary holding parameters from the Excel sheet
    :rtype: dict
    """
    # PV2AC conversion pathway TODO
    if parameter['Top'] == 'DC' or parameter['Top'] == 'PVINV' or parameter['Top'] == 'PV' and parameter['P_PV2AC_out'] is not None or parameter['Top'] == 'AC' and parameter['P_PV2AC_out'] is not None:
        
        # Create variables for the sampling points and corresponding efficiencies TODO
        p_pv2ac = np.fromiter((value for key, value in parameter.items() if 'p_PV2AC_' in key and value is not None), float)
        eta_pv2ac = np.fromiter((value / 100 for key, value in parameter.items() if 'eta_PV2AC_' in key and value is not None), float)

        # Absolute input and output power in W
        p_pv2ac_out = parameter['P_PV2AC_out'] * p_pv2ac * 1000
        p_pv2ac_in = p_pv2ac_out / eta_pv2ac

        # Absolute power loss in W
        P_l_pv2ac_in = (1 - eta_pv2ac) * p_pv2ac_in
        P_l_pv2ac_out = (1 / eta_pv2ac - 1) * p_pv2ac_out

        # Polynomial curve fitting parameters of the power loss functions in W
        
        # Based on input power
        p = np.polyfit(p_pv2ac_in / parameter['P_PV2AC_in'] / 1000, P_l_pv2ac_in, 2)
        parameter['PV2AC_a_in'] = p[0]
        parameter['PV2AC_b_in'] = p[1]
        parameter['PV2AC_c_in'] = p[2]

        # Based on output power
        p = np.polyfit(p_pv2ac, P_l_pv2ac_out, 2)
        parameter['PV2AC_a_out'] = p[0]
        parameter['PV2AC_b_out'] = p[1]
        parameter['PV2AC_c_out'] = p[2]
    
    # PV2BAT conversion pathway
    if parameter['Top'] == 'DC' or parameter['Top'] == 'PV':

        # Create variables for the sampling points and corresponding efficiencies
        p_pv2bat = np.array([value for key, value in parameter.items() if 'p_PV2BAT_' in key])
        eta_pv2bat = np.array([value / 100 for key, value in parameter.items() if 'eta_PV2BAT_' in key])

        # Create missing variables

        # Nominal input power of the PV2BAT conversion pathway of DC-coupled systems
        if parameter['P_PV2BAT_in'] is None:
            parameter['P_PV2BAT_in'] = parameter['P_PV2BAT_out'] / (parameter['eta_PV2BAT_100'] / 100)

        # Absolute input and output power in W
        p_pv2bat_out = parameter['P_PV2BAT_out'] * p_pv2bat * 1000
        p_pv2bat_in = p_pv2bat_out / eta_pv2bat

        # Absolute power loss in W
        P_l_pv2bat_in = (1 - eta_pv2bat) * p_pv2bat_in
        P_l_pv2bat_out = (1 / eta_pv2bat - 1) * p_pv2bat_out
        
        # Polynomial curve fitting parameters of the power loss functions in W
                 
        # Based on input power
        p = np.polyfit(p_pv2bat_in / parameter['P_PV2BAT_in'] / 1000, P_l_pv2bat_in, 2)
        parameter['PV2BAT_a_in'] = p[0]
        parameter['PV2BAT_b_in'] = p[1]
        parameter['PV2BAT_c_in'] = p[2]

        # Based on output power
        p = np.polyfit(p_pv2bat, P_l_pv2bat_out, 2)
        parameter['PV2BAT_a_out'] = p[0]
        parameter['PV2BAT_b_out'] = p[1]
        parameter['PV2BAT_c_out'] = p[2]
    
    # AC2BAT conversion pathway
    if parameter['Top'] == 'AC' or parameter['Top'] == 'DC' and parameter['P_AC2BAT_in'] is not None:

        # Create variables for the sampling points and corresponding efficiencies TODO
        p_ac2bat = np.fromiter((value for key, value in parameter.items() if 'p_AC2BAT_' in key), float)
        eta_ac2bat = np.fromiter((value / 100 for key, value in parameter.items() if 'eta_AC2BAT_' in key), float)

        # Absolute input and output power in W
        p_ac2bat_out = parameter['P_PV2BAT_out'] * p_ac2bat * 1000
        p_ac2bat_in = p_ac2bat_out / eta_ac2bat

        # Absolute power loss in W
        P_l_ac2bat_in = (1 - eta_ac2bat) * p_ac2bat_in
        P_l_ac2bat_out = (1 / eta_ac2bat - 1) * p_ac2bat_out

        # Polynomial curve fitting parameters of the power loss functions in W
        
        # Based on input power
        p = np.polyfit(p_ac2bat_in / parameter['P_AC2BAT_in'] / 1000, P_l_ac2bat_in, 2)
        parameter['AC2BAT_a_in'] = p[0]
        parameter['AC2BAT_b_in'] = p[1]
        parameter['AC2BAT_c_in'] = p[2]

        # Based on output power
        p = np.polyfit(p_ac2bat, P_l_ac2bat_out, 2)
        parameter['AC2BAT_a_out'] = p[0]
        parameter['AC2BAT_b_out'] = p[1]
        parameter['AC2BAT_c_out'] = p[2]
    
    # BAT2AC conversion pathway
    if parameter['Top'] =='AC' or parameter['Top'] =='DC' or parameter['Top'] =='PV' and parameter['P_BAT2AC_out'] is not None:

        # Create variables for the sampling points and corresponding efficiencies TODO
        p_bat2ac = np.fromiter((value for key, value in parameter.items() if 'p_BAT2AC_' in key), float)
        eta_bat2ac = np.fromiter((value / 100 for key, value in parameter.items() if 'eta_BAT2AC_' in key), float)

        # Absolute input and output power in W
        p_bat2ac_out = parameter['P_BAT2AC_out'] * p_bat2ac * 1000
        p_bat2ac_in = p_bat2ac_out / eta_bat2ac

        # Absolute power loss in W
        P_l_bat2ac_in = (1 - eta_bat2ac) * p_bat2ac_in
        P_l_bat2ac_out = (1 / eta_bat2ac - 1) * p_bat2ac_out

        # Polynomial curve fitting parameters of the power loss functions in W
        
        # Based on input power
        p = np.polyfit(p_bat2ac_in / parameter['P_BAT2AC_in'] / 1000, P_l_bat2ac_in, 2)
        parameter['BAT2AC_a_in'] = p[0]
        parameter['BAT2AC_b_in'] = p[1]
        parameter['BAT2AC_c_in'] = p[2]

        # Based on output power
        p = np.polyfit(p_bat2ac, P_l_bat2ac_out, 2)
        parameter['BAT2AC_a_out'] = p[0]
        parameter['BAT2AC_b_out'] = p[1]
        parameter['BAT2AC_c_out'] = p[2]
    
    # BAT2PV conversion pathway
    if parameter['Top'] =='PV':

        # Create variables for the sampling points and corresponding efficiencies TODO
        p_bat2pv = np.fromiter((value for key, value in parameter.items() if 'p_BAT2PV_' in key), float)
        eta_bat2pv = np.fromiter((value / 100 for key, value in parameter.items() if 'eta_BAT2PV_' in key), float)

        # Absolute input and output power in W
        p_bat2pv_out = parameter['P_BAT2PV_out'] * p_bat2pv * 1000
        p_bat2pv_in = p_bat2pv_out / eta_bat2pv

        # Absolute power loss in W
        P_l_bat2pv_in = (1 - eta_bat2pv) * p_bat2pv_in
        P_l_bat2pv_out = (1 / eta_bat2pv - 1) * p_bat2pv_out

        # Polynomial curve fitting parameters of the power loss functions in W
        
        # Based on input power TODO
        p = np.polyfit(p_bat2pv_in / parameter['P_BAT2AC_in'] / 1000, P_l_bat2pv_in, 2)
        parameter['BAT2PV_a_in'] = p[0]
        parameter['BAT2PV_b_in'] = p[1]
        parameter['BAT2PV_c_in'] = p[2]

        # Based on output power
        p = np.polyfit(p_bat2pv, P_l_bat2pv_out, 2)
        parameter['BAT2PV_a_out'] = p[0]
        parameter['BAT2PV_b_out'] = p[1]
        parameter['BAT2PV_c_out'] = p[2]
    
    # Additional parameters

    # Mean battery capacity in kWh
    try:
        parameter['E_BAT'] = (parameter['E_BAT_usable'] / parameter['eta_BAT'] * 100 + parameter['E_BAT_usable']) / 2
    except:
        parameter['E_BAT'] = None

    # Mean stationary deviation of the charging power in W
    try:
        parameter['P_PV2BAT_DEV'] = parameter['P_PV2BAT_DEV_IMPORT'] - parameter['P_PV2BAT_DEV_EXPORT']
    except:
        parameter['P_PV2BAT_DEV'] = None

    if parameter['Top'] == 'AC':
        parameter['P_AC2BAT_DEV'] = parameter['P_PV2BAT_DEV'] 
    
    # Mean stationary deviation of the discharging power in W
    try:
        parameter['P_BAT2AC_DEV'] = parameter['P_BAT2AC_DEV_EXPORT'] - parameter['P_BAT2AC_DEV_IMPORT']
    except:
        parameter['P_BAT2AC_DEV'] = None
    
    # Time constant for the first-order time delay element in s
    try:
        parameter['t_CONSTANT'] = (parameter['t_SETTLING'] - round(parameter['t_DEAD'])) / 3
    except:
        parameter['t_CONSTANT'] = None

    # Hysteresis threshold for the recharging of the battery
    parameter['SOC_h'] = 0.98

    # Feed-in power limit in kW/kWp
    parameter['p_ac2g_max'] = 0.7

    return parameter

def load_mat(fname, name):
    """Loads mat files

    :param fname: Path to mat file 
    :type fname: string
    :return: Data from the mat file
    :rtype: numpy array
    """
    _version = '0.1'
    # Loads content of the mat-file
    mat_contents = sio.loadmat(fname, squeeze_me=True)
    # extracts specific data
    data = np.array(mat_contents[name], dtype='float64')
    
    return data 

def load_ref_case(fname, name):

    with open(fname, 'rb') as f:

        a = np.load(f)

        data = a[name]
    
    return data

def resample_data_frame(df):
    """Function for resampling data frames

    :param df: data frame
    :type df: pandas data frame
    :return: data frame
    :rtype: pandas data frame
    """
    df_rs = df.resample('15min').mean()

    return df_rs

def transform_dict_to_array(parameter):
    """Function for transforming a dict to an numpy array

    :param parameter: dict of system parameters
    :type parameter: dict
    :return: array of system parameters
    :rtype: numpy array
    """
    if parameter['Top'] == 'AC':
        d = np.array(parameter['E_BAT'])                # 0
        d = np.append(d, parameter['eta_BAT'])          # 1
        d = np.append(d, parameter['t_CONSTANT'])  # 2
        d = np.append(d, parameter['P_SYS_SOC0_DC'])  # 3
        d = np.append(d, parameter['P_SYS_SOC0_AC'])  # 4
        d = np.append(d, parameter['P_SYS_SOC1_DC'])  # 5
        d = np.append(d, parameter['P_SYS_SOC1_AC'])  # 6
        d = np.append(d, parameter['AC2BAT_a_in'])  # 7
        d = np.append(d, parameter['AC2BAT_b_in'])  # 8
        d = np.append(d, parameter['AC2BAT_c_in'])  # 9
        d = np.append(d, parameter['BAT2AC_a_out'])  # 10
        d = np.append(d, parameter['BAT2AC_b_out'])  # 11
        d = np.append(d, parameter['BAT2AC_c_out'])  # 12
        d = np.append(d, parameter['P_AC2BAT_DEV'])  # 13
        d = np.append(d, parameter['P_BAT2AC_DEV'])  # 14
        d = np.append(d, parameter['P_BAT2AC_out'])  # 15
        d = np.append(d, parameter['P_AC2BAT_in'])  # 16
        d = np.append(d, parameter['t_DEAD'])  # 17
        d = np.append(d, parameter['SOC_h'])  # 18

    if parameter['Top'] == 'DC':
        d = np.array(parameter['E_BAT'])  # 1
        d = np.append(d, parameter['P_PV2AC_in'])  # 2
        d = np.append(d, parameter['P_PV2AC_out'])  # 3
        d = np.append(d, parameter['P_PV2BAT_in'])       # 4
        d = np.append(d, parameter['P_BAT2AC_out'])       # 5
        d = np.append(d, parameter['PV2AC_a_in'])          # 6
        d = np.append(d, parameter['PV2AC_b_in'])           # 7
        d = np.append(d, parameter['PV2AC_c_in'])            # 8
        d = np.append(d, parameter['PV2BAT_a_in'])            # 9
        d = np.append(d, parameter['PV2BAT_b_in'])  # 10
        d = np.append(d, parameter['BAT2AC_a_out'])  # 11
        d = np.append(d, parameter['BAT2AC_b_out'])  # 12
        d = np.append(d, parameter['BAT2AC_c_out'])  # 13
        d = np.append(d, parameter['eta_BAT'])  # 14
        d = np.append(d, parameter['SOC_h'])  # 15
        d = np.append(d, parameter['P_PV2BAT_DEV'])       # 16
        d = np.append(d, parameter['P_BAT2AC_DEV'])  # 17
        d = np.append(d, parameter['t_DEAD'])  # 18
        d = np.append(d, parameter['t_CONSTANT'])          # 19
        d = np.append(d, parameter['P_SYS_SOC1_DC'])  # 20
        d = np.append(d, parameter['P_SYS_SOC0_AC'])  # 21
        d = np.append(d, parameter['P_SYS_SOC0_DC'])  # 22

    if parameter['Top'] == 'PV':
        d = np.array(parameter['E_BAT'])
        d = np.append(d, parameter['P_PV2AC_in'])
        d = np.append(d, parameter['P_PV2AC_out'])
        d = np.append(d, parameter['P_PV2BAT_in'])
        d = np.append(d, parameter['P_BAT2PV_out'])
        d = np.append(d, parameter['PV2AC_a_in'])
        d = np.append(d, parameter['PV2AC_b_in'])
        d = np.append(d, parameter['PV2AC_c_in'])
        d = np.append(d, parameter['PV2BAT_a_in'])
        d = np.append(d, parameter['PV2BAT_b_in'])
        d = np.append(d, parameter['PV2BAT_c_in'])
        d = np.append(d, parameter['PV2AC_a_out'])
        d = np.append(d, parameter['PV2AC_b_out'])
        d = np.append(d, parameter['PV2AC_c_out'])
        d = np.append(d, parameter['BAT2PV_a_out'])
        d = np.append(d, parameter['BAT2PV_b_out'])
        d = np.append(d, parameter['BAT2PV_c_out'])
        d = np.append(d, parameter['eta_BAT'])
        d = np.append(d, parameter['SOC_h'])
        d = np.append(d, parameter['P_PV2BAT_DEV'])
        d = np.append(d, parameter['P_BAT2AC_DEV'])
        d = np.append(d, parameter['P_SYS_SOC1_DC'])
        d = np.append(d, parameter['P_SYS_SOC0_AC'])
        d = np.append(d, parameter['P_SYS_SOC0_DC'])
        d = np.append(d, parameter['t_DEAD'])
        d = np.append(d, parameter['t_CONSTANT'])

    return d

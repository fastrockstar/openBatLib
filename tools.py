"""
Collection of functions used in BatMod
TODO:
    BAT2PV conversion pathway polyfit mit P_BAT2AC_in?
    Umwnadlung in MWh dt berücksichitge
"""
import scipy.io as sio
import numpy as np
from openpyxl import load_workbook
import numba as nb
import time

def load_mat(fname, name):
    """Loads mat files

    :param fname: Path to mat file 
    :type fname: string
    :return: Data from the mat file
    :rtype: numpy array
    """
    _version = '0.1'
    # Loads content of the mat-file
    mat_contents = sio.loadmat(fname, squeeze_me=True)
    # extracts specific data
    data = np.array(mat_contents[name], dtype='float64')
    
    return data 

def load_parameter(fname, col_name):
    """Loads system parameter from excel file

    :param fname: Path to the excel file
    :type fname: string
    :param col_name: Column to read data from
    :type col_name: string
    :return: Dictionary holding parameters from the Excel sheet
    :rtype: dict
    """
    _version = '0.1'
    
    wb = load_workbook(fname, data_only=True)
    ws = wb['Data'] # Load Data sheet of excel file

    # read keys and values from Excel sheet
    keys = (c.value for c in ws['E'][1:])
    values = (c.value if c.value != 'ns' else None for c in ws[col_name][1:])
    
    parameter = dict(zip(keys, values))

    # deletes entries where key is None
    del parameter[None]

    # Assign specific parameters TODO
    parameter['P_PV2AC_out_PVINV'] = ws[col_name][15].value
    parameter['P_PV2AC_out'] = ws[col_name][24].value
    parameter['P_AC2BAT_in_DCC'] = ws[col_name][25].value
    parameter['P_AC2BAT_in'] = ws[col_name][26].value
    parameter['P_BAT2AC_out'] = ws[col_name][27].value
    parameter['P_BAT2AC_out_DCC'] = ws[col_name][28].value

    # Set refrence case values to boolean
    if parameter['ref_1'] == 'yes':
        parameter['ref_1'] = True
    elif parameter['ref_1'] == 'no':
        parameter['ref_1'] = False
    
    if parameter['ref_2'] == 'yes':
        parameter['ref_2'] = True
    elif parameter['ref_2'] == 'no':
        parameter['ref_2'] = False

    # Specific parameters of DC-coupled systems
    if parameter['Top'] == 'DC':
        parameter['P_AC2BAT_in'] = parameter['P_AC2BAT_in_DCC'] # Nominal charging power (AC) in kW
        parameter['P_BAT2AC_out'] = parameter['P_BAT2AC_out_DCC']
    
    # Specific parameters of PV inverters and AC-coupled systems SIEHE TODO
    if parameter['Top'] == 'PVINV' or parameter['Top'] == 'AC' and parameter['P_PV2AC_out_PVINV'] is not None:
        parameter['P_PV2AC_out'] = parameter['P_PV2AC_out_PVINV']
    
    # Specific parameters of PV-coupled systems
    if parameter['Top'] == 'PV':
        parameter['P_BAT2PV'] == parameter['P_BAT2AC_in']
        parameter['P_BAT2AC_out'] = parameter['P_BAT2AC_out_DCC']

    # replace 'ns', 'o' and 'c' entries to None
    for key, value in parameter.items():
        if value == 'ns' or value == 'o' or value == 'c' or value == ' ':
            parameter[key] = None

    # Convert to kW
    convert_to_kw = ['P_PV2AC_in', 'P_PV2AC_out_PVINV','P_PV2AC_out','P_AC2BAT_in_DCC','P_AC2BAT_in','P_BAT2AC_out',
             'P_BAT2AC_out_DCC','P_PV2BAT_in','P_BAT2PV_out','P_PV2BAT_out','P_BAT2AC_in']

    for par in convert_to_kw:
        if parameter[par]:
            parameter[par] /= 1000
        
    return parameter

def eta2abc(parameter):
    """Function to calculate the parameters of the power loss functions (quadratic equations) from the path efficiencies

    :param parameter: Holds parameters of the system
    :type parameter: dict
    :return: Dictionary holding parameters from the Excel sheet
    :rtype: dict
    """
    # PV2AC conversion pathway TODO
    if parameter['Top'] == 'DC' or parameter['Top'] == 'PVINV' or parameter['Top'] == 'PV' and parameter['P_PV2AC_out'] is not None or parameter['Top'] == 'AC' and parameter['P_PV2AC_out'] is not None:
        
        # Create variables for the sampling points and corresponding efficiencies TODO
        p_pv2ac = np.fromiter((value for key, value in parameter.items() if 'p_PV2AC_' in key and value is not None), float)
        eta_pv2ac = np.fromiter((value / 100 for key, value in parameter.items() if 'eta_PV2AC_' in key and value is not None), float)

        # Absolute input and output power in W
        p_pv2ac_out = parameter['P_PV2AC_out'] * p_pv2ac * 1000
        p_pv2ac_in = p_pv2ac_out / eta_pv2ac

        # Absolute power loss in W
        P_l_pv2ac_in = (1 - eta_pv2ac) * p_pv2ac_in
        P_l_pv2ac_out = (1 / eta_pv2ac - 1) * p_pv2ac_out

        # Polynomial curve fitting parameters of the power loss functions in W
        
        # Based on input power
        p = np.polyfit(p_pv2ac_in / parameter['P_PV2AC_in'] / 1000, P_l_pv2ac_in, 2)
        parameter['PV2AC_a_in'] = p[0]
        parameter['PV2AC_b_in'] = p[1]
        parameter['PV2AC_c_in'] = p[2]

        # Based on output power
        p = np.polyfit(p_pv2ac, P_l_pv2ac_out, 2)
        parameter['PV2AC_a_out'] = p[0]
        parameter['PV2AC_b_out'] = p[1]
        parameter['PV2AC_c_out'] = p[2]
    
    # PV2BAT conversion pathway
    if parameter['Top'] == 'DC' or parameter['Top'] == 'PV':

        # Create variables for the sampling points and corresponding efficiencies
        p_pv2bat = np.array([value for key, value in parameter.items() if 'p_PV2BAT_' in key])
        eta_pv2bat = np.array([value / 100 for key, value in parameter.items() if 'eta_PV2BAT_' in key])

        # Create missing variables

        # Nominal input power of the PV2BAT conversion pathway of DC-coupled systems
        if parameter['P_PV2BAT_in'] is None:
            parameter['P_PV2BAT_in'] = parameter['P_PV2BAT_out'] / (parameter['eta_PV2BAT_100'] / 100)

        # Absolute input and output power in W
        p_pv2bat_out = parameter['P_PV2BAT_out'] * p_pv2bat * 1000
        p_pv2bat_in = p_pv2bat_out / eta_pv2bat

        # Absolute power loss in W
        P_l_pv2bat_in = (1 - eta_pv2bat) * p_pv2bat_in
        P_l_pv2bat_out = (1 / eta_pv2bat - 1) * p_pv2bat_out
        
        # Polynomial curve fitting parameters of the power loss functions in W
                 
        # Based on input power
        p = np.polyfit(p_pv2bat_in / parameter['P_PV2BAT_in'] / 1000, P_l_pv2bat_in, 2)
        parameter['PV2BAT_a_in'] = p[0]
        parameter['PV2BAT_b_in'] = p[1]
        parameter['PV2BAT_c_in'] = p[2]

        # Based on output power
        p = np.polyfit(p_pv2bat, P_l_pv2bat_out, 2)
        parameter['PV2BAT_a_out'] = p[0]
        parameter['PV2BAT_b_out'] = p[1]
        parameter['PV2BAT_c_out'] = p[2]
    
    # AC2BAT conversion pathway
    if parameter['Top'] == 'AC' or parameter['Top'] == 'DC' and parameter['P_AC2BAT_in'] is not None:

        # Create variables for the sampling points and corresponding efficiencies TODO
        p_ac2bat = np.fromiter((value for key, value in parameter.items() if 'p_AC2BAT_' in key), float)
        eta_ac2bat = np.fromiter((value / 100 for key, value in parameter.items() if 'eta_AC2BAT_' in key), float)

        # Absolute input and output power in W
        p_ac2bat_out = parameter['P_PV2BAT_out'] * p_ac2bat * 1000
        p_ac2bat_in = p_ac2bat_out / eta_ac2bat

        # Absolute power loss in W
        P_l_ac2bat_in = (1 - eta_ac2bat) * p_ac2bat_in
        P_l_ac2bat_out = (1 / eta_ac2bat - 1) * p_ac2bat_out

        # Polynomial curve fitting parameters of the power loss functions in W
        
        # Based on input power
        p = np.polyfit(p_ac2bat_in / parameter['P_AC2BAT_in'] / 1000, P_l_ac2bat_in, 2)
        parameter['AC2BAT_a_in'] = p[0]
        parameter['AC2BAT_b_in'] = p[1]
        parameter['AC2BAT_c_in'] = p[2]

        # Based on output power
        p = np.polyfit(p_ac2bat, P_l_ac2bat_out, 2)
        parameter['AC2BAT_a_out'] = p[0]
        parameter['AC2BAT_b_out'] = p[1]
        parameter['AC2BAT_c_out'] = p[2]
    
    # BAT2AC conversion pathway
    if parameter['Top'] =='AC' or parameter['Top'] =='DC' or parameter['Top'] =='PV' and parameter['P_BAT2AC_out'] is not None:

        # Create variables for the sampling points and corresponding efficiencies TODO
        p_bat2ac = np.fromiter((value for key, value in parameter.items() if 'p_BAT2AC_' in key), float)
        eta_bat2ac = np.fromiter((value / 100 for key, value in parameter.items() if 'eta_BAT2AC_' in key), float)

        # Absolute input and output power in W
        p_bat2ac_out = parameter['P_BAT2AC_out'] * p_bat2ac * 1000
        p_bat2ac_in = p_bat2ac_out / eta_bat2ac

        # Absolute power loss in W
        P_l_bat2ac_in = (1 - eta_bat2ac) * p_bat2ac_in
        P_l_bat2ac_out = (1 / eta_bat2ac - 1) * p_bat2ac_out

        # Polynomial curve fitting parameters of the power loss functions in W
        
        # Based on input power
        p = np.polyfit(p_bat2ac_in / parameter['P_BAT2AC_in'] / 1000, P_l_bat2ac_in, 2)
        parameter['BAT2AC_a_in'] = p[0]
        parameter['BAT2AC_b_in'] = p[1]
        parameter['BAT2AC_c_in'] = p[2]

        # Based on output power
        p = np.polyfit(p_bat2ac, P_l_bat2ac_out, 2)
        parameter['BAT2AC_a_out'] = p[0]
        parameter['BAT2AC_b_out'] = p[1]
        parameter['BAT2AC_c_out'] = p[2]
    
    # BAT2PV conversion pathway
    if parameter['Top'] =='PV':

        # Create variables for the sampling points and corresponding efficiencies TODO
        p_bat2pv = np.fromiter((value for key, value in parameter.items() if 'p_BAT2PV_' in key), float)
        eta_bat2pv = np.fromiter((value / 100 for key, value in parameter.items() if 'eta_BAT2PV_' in key), float)

        # Absolute input and output power in W
        p_bat2pv_out = parameter['P_BAT2PV_out'] * p_bat2pv * 1000
        p_bat2pv_in = p_bat2pv_out / eta_bat2pv

        # Absolute power loss in W
        P_l_bat2pv_in = (1 - eta_bat2pv) * p_bat2pv_in
        P_l_bat2pv_out = (1 / eta_bat2pv - 1) * p_bat2pv_out

        # Polynomial curve fitting parameters of the power loss functions in W
        
        # Based on input power TODO
        p = np.polyfit(p_bat2pv_in / parameter['P_BAT2AC_in'] / 1000, P_l_bat2pv_in, 2)
        parameter['BAT2PV_a_in'] = p[0]
        parameter['BAT2PV_b_in'] = p[1]
        parameter['BAT2PV_c_in'] = p[2]

        # Based on output power
        p = np.polyfit(p_bat2ac, P_l_bat2pv_out, 2)
        parameter['BAT2PV_a_out'] = p[0]
        parameter['BAT2PV_b_out'] = p[1]
        parameter['BAT2PV_c_out'] = p[2]
    
    # Additional parameters

    # Mean battery capacity in kWh
    try:
        parameter['E_BAT'] = (parameter['E_BAT_usable'] / parameter['eta_BAT'] * 100 + parameter['E_BAT_usable']) / 2
    except:
        parameter['E_BAT'] = None

    # Mean stationary deviation of the charging power in W
    try:
        parameter['P_PV2BAT_DEV'] = parameter['P_PV2BAT_DEV_IMPORT'] - parameter['P_PV2BAT_DEV_EXPORT']
    except:
        parameter['P_PV2BAT_DEV'] = None

    if parameter['Top'] == 'AC':
        parameter['P_AC2BAT_DEV'] = parameter['P_PV2BAT_DEV'] 
    
    # Mean stationary deviation of the discharging power in W
    try:
        parameter['P_BAT2AC_DEV'] = parameter['P_BAT2AC_DEV_EXPORT'] - parameter['P_BAT2AC_DEV_IMPORT']
    except:
        parameter['P_BAT2AC_DEV'] = None
    
    # Time constant for the first-order time delay element in s
    try:
        parameter['t_CONSTANT'] = (parameter['t_SETTLING'] - round(parameter['t_DEAD'])) / 3
    except:
        parameter['t_CONSTANT'] = None

    # Hysteresis threshold for the recharging of the battery
    parameter['SOC_h'] = 0.98

    # Feed-in power limit in kW/kWp
    parameter['p_ac2g_max'] = 0.7

    return parameter

@nb.jit(nopython=True)
def run_ideal_DC(lenght, _soc0, _E_BAT, _Pr, _dt, _Pbat, _soc):

    _E_BAT *= 1000
    _dt /= 3600

    for i in range(lenght):
        # Energy content of the battery in the previous time step
        E_b0 = _soc0 * _E_BAT

        # Calculate the DC power of the battery from the residual power
        P_bat = _Pr[i]

        # Decision whether the battery should be charged or discharged
        if P_bat > 0 and _soc0 < 1:
            # Change the energy content of the battery
            E_b = E_b0 + P_bat * _dt
        elif P_bat < 0 and _soc0 > 0:
            E_b = E_b0 + P_bat * _dt
        
        # Neither charging nor discharging of the battery
        else:
            P_bat = 0
            E_b = E_b0

        # Transfer the realized DC power of the battery
        _Pbat[i] = P_bat

        # Calculate the state of charge of the battery
        _soc0 = E_b / _E_BAT
        _soc[i] = _soc0
    
    return _soc0, _Pbat, _soc

@nb.jit(nopython=True)
def run_loss_DC(_E_BAT, _P_PV2AC_in, _P_PV2AC_out, _P_PV2BAT_in, _P_BAT2AC_out, _PV2AC_a_in, _PV2AC_b_in, _PV2AC_c_in, _PV2BAT_a_in, _PV2BAT_b_in, _BAT2AC_a_out, _BAT2AC_b_out, _BAT2AC_c_out, _eta_BAT, _SOC_h, _P_PV2BAT_DEV, _P_BAT2AC_DEV, _t_DEAD, _t_CONSTANT, _P_SYS_SOC1_DC, _P_SYS_SOC0_AC, _P_SYS_SOC0_DC, _P_PV2AC_min, _tend, _soc0, _Prpv, _Pr, _Ppv, _Ppv2bat_in, _ppv2ac, _Ppv2ac_out, _Pbat2ac_out, _Ppvbs, _Pbat, _soc, _dt, _th):
    '''
    TODO 1 t_start auf das 2. Element?
    '''
    _E_BAT *= 1000
    _P_PV2AC_out *= 1000
    _eta_BAT /= 100

    # Initialization of particular variables
    #_P_PV2AC_min = _parameter['PV2AC_c_in'] # Minimum input power of the PV2AC conversion pathway
    _tde = _t_CONSTANT > 0 # Binary variable to activate the first-order time delay element
    _ftde = 1 - np.exp(-_dt / _t_CONSTANT) # Factor of the first-order time delay element
    _tstart = np.maximum(2, 1 + _t_DEAD) # First time step with regard to the dead time of the system control
    korr = 0.1

    if _dt >= (3 * _t_CONSTANT):
        _tstart = 1
        T_DEAD = False
    else:
        T_DEAD = True

    if _dt >= _t_DEAD + 3 * _t_CONSTANT:
        SETTLING = False
    else:
        SETTLING = True


    #for t in range(_tstart - 1, _tend):
    for t in range(0, _tend):
        # Energy content of the battery in the previous time step
        E_b0 = _soc0 * _E_BAT

        # Residual power with regard to the dead time of the system control
        
        if T_DEAD:
            P_rpv = _Prpv[t - _t_DEAD] 
            P_r = _Pr[t - _t_DEAD]

        else:
            P_rpv = _Prpv[t] 
            P_r = _Pr[t]

        # Check if the battery holds enough unused capacity for charging or discharging
        # Estimated amount of energy that is supplied to or discharged from the storage unit.
        E_bs_rpv = P_rpv * _dt / 1000
        E_bs_r = P_r * _dt / 1000
        
        if E_bs_rpv > 0 and E_bs_rpv > (_E_BAT - E_b0):
            P_rpv = (_E_BAT - E_b0) / _dt
        # wenn Laden, dann neue Ladeleistung inkl. Korrekturfaktor
        elif E_bs_r < 0 and np.abs(E_bs_r) > (E_b0):
            P_r = (E_b0) / _dt * (1-korr)
        
        # Decision if the battery should be charged or discharged
        if P_rpv > 0 and _soc0 < 1 - _th * (1 - _SOC_h):
            '''
            The last term th*(1-SOC_h) avoids the alternation between
            charging and standby mode due to the DC power consumption of the
            battery converter when the battery is fully charged. The battery
            will not be recharged until the SOC falls below the SOC-threshold
            (SOC_h) for recharging from PV.
            '''
            # Charging power
            P_pv2bat_in = P_rpv
            
            # Adjust the charging power due to the stationary deviations
            P_pv2bat_in = np.maximum(0, P_pv2bat_in + _P_PV2BAT_DEV)
            
            # Limit the charging power to the maximum charging power
            P_pv2bat_in = np.minimum(P_pv2bat_in, _P_PV2BAT_in * 1000)
            
            # Adjust the charging power due to the settling time
            # (modeled by a first-order time delay element)
            if SETTLING:
                P_pv2bat_in = _tde * _Ppv2bat_in[(t-1)] + _tde * (P_pv2bat_in - _Ppv2bat_in[(t-1)]) * _ftde + P_pv2bat_in * (not _tde)
            
            # Limit the charging power to the current power output of the PV generator
            P_pv2bat_in = np.minimum(P_pv2bat_in, _Ppv[t])
            
            # Normalized charging power
            ppv2bat = P_pv2bat_in / _P_PV2BAT_in / 1000
            
            # DC power of the battery affected by the PV2BAT conversion losses
            # (the idle losses of the PV2BAT conversion pathway are not taken
            # into account)
            P_bat = np.maximum(0, P_pv2bat_in - (_PV2BAT_a_in * ppv2bat**2 + _PV2BAT_b_in * ppv2bat))
            
            # Realized DC input power of the PV2AC conversion pathway
            P_pv2ac_in = _Ppv[t] - P_pv2bat_in
            
            # Normalized DC input power of the PV2AC conversion pathway
            _ppv2ac = P_pv2ac_in / _P_PV2AC_in / 1000
            
            # Realized AC power of the PV-battery system
            P_pv2ac_out = np.maximum(0, P_pv2ac_in - (_PV2AC_a_in * _ppv2ac**2 + _PV2AC_b_in * _ppv2ac + _PV2AC_c_in))
            P_pvbs = P_pv2ac_out
            
            # Transfer the final values
            _Ppv2ac_out[t] = P_pv2ac_out
            _Ppv2bat_in[t] = P_pv2bat_in

        elif P_rpv < 0 and _soc0 > 0: 

            # Discharging power
            P_bat2ac_out = P_r * -1
            
            # Adjust the discharging power due to the stationary deviations
            P_bat2ac_out = np.maximum(0, P_bat2ac_out + _P_BAT2AC_DEV)
            
            # Adjust the discharging power to the maximum discharging power
            P_bat2ac_out = np.minimum(P_bat2ac_out, _P_BAT2AC_out * 1000)
            
            # Adjust the discharging power due to the settling time
            # (modeled by a first-order time delay element)
            if SETTLING:
                P_bat2ac_out = _tde * _Pbat2ac_out[t-1] + _tde * (P_bat2ac_out - _Pbat2ac_out[t-1]) * _ftde + P_bat2ac_out * (not _tde)
            
            # Limit the discharging power to the maximum AC power output of the PV-battery system
            P_bat2ac_out = np.minimum(_P_PV2AC_out - _Ppv2ac_out[t], P_bat2ac_out)
            
            # Normalized discharging power
            ppv2bat = P_bat2ac_out / _P_BAT2AC_out / 1000
            
            # DC power of the battery affected by the BAT2AC conversion losses
            # (if the idle losses of the PV2AC conversion pathway are covered by
            # the PV generator, the idle losses of the BAT2AC conversion pathway
            # are not taken into account)
            if _Ppv[t] > _P_PV2AC_min:
                P_bat= -1 * (P_bat2ac_out + (_BAT2AC_a_out * ppv2bat**2 + _BAT2AC_b_out * ppv2bat))
            else:
                P_bat = -1 * (P_bat2ac_out + (_BAT2AC_a_out * ppv2bat**2 + _BAT2AC_b_out * ppv2bat + _BAT2AC_c_out)) + _Ppv[t]
            
                    
            # Realized AC power of the PV-battery system
            P_pvbs = _Ppv2ac_out[t] + P_bat2ac_out
            
            # Transfer the final values
            _Pbat2ac_out[t] = P_bat2ac_out

        else: # Neither charging nor discharging of the battery

            # Set the DC power of the battery to zero
            P_bat = 0
        
            # Realized AC power of the PV-battery system
            P_pvbs = _Ppv2ac_out[t]

        
        # Decision if the standby mode is active
        if P_bat == 0 and P_pvbs == 0 and _soc0 <= 0: # Standby mode in discharged state

            # DC and AC power consumption of the PV-battery inverter
            P_bat = -np.maximum(0, _P_SYS_SOC0_DC)
            P_pvbs = -_P_SYS_SOC0_AC

        elif P_bat == 0 and P_pvbs > 0 and _soc0 > 0: # Standby mode in fully charged state

            # DC power consumption of the PV-battery inverter
            P_bat = -np.maximum(0, _P_SYS_SOC1_DC)

        
        # Transfer the realized AC power of the PV-battery system and the DC power of the battery
        _Ppvbs[t] = P_pvbs
        _Pbat[t] = P_bat
        
        # Change the energy content of the battery Wx to Wh conversion
        if P_bat > 0:
            E_b = E_b0 + P_bat * np.sqrt(_eta_BAT) * _dt / 3600
        elif P_bat < 0:
            E_b = E_b0 + P_bat / np.sqrt(_eta_BAT) * _dt / 3600
        else:
            E_b = E_b0
                    
        # Calculate the state of charge of the battery
        _soc0 = E_b / _E_BAT
        _soc[t] = _soc0

        # Adjust the hysteresis threshold to avoid alternation between charging
        # and standby mode due to the DC power consumption of the
        # PV-battery inverter
        if _th and _soc[t] > _SOC_h or _soc[t] > 1:
            _th = True
        else:
            _th = False 

    return _Ppv2ac_out, _Ppv2bat_in, _Pbat2ac_out, _Ppvbs, _Pbat, _soc, _soc0

@nb.jit(nopython=True)
def run_loss_AC(_E_BAT, _eta_BAT, _t_CONSTANT, _P_SYS_SOC0_DC, _P_SYS_SOC0_AC, _P_SYS_SOC1_DC, _P_SYS_SOC1_AC, _AC2BAT_a_in, _AC2BAT_b_in, _AC2BAT_c_in, _BAT2AC_a_out, _BAT2AC_b_out, _BAT2AC_c_out, _P_AC2BAT_DEV, _P_BAT2AC_DEV, _P_BAT2AC_out, _P_AC2BAT_in, _t_DEAD , _SOC_h, _dt, _th, _soc0, _tend, _soc, _Pr, _Pbs, _Pbat):
    # Initialization of particular variables
    _P_AC2BAT_min = _AC2BAT_c_in # Minimum AC charging power
    _P_BAT2AC_min = _BAT2AC_c_out # Minimum AC discharging power

    # Correction factor to avoid over charge and discharge the battery
    corr = 0.1

    # Initialization of particular variables
    #_P_PV2AC_min = _parameter['PV2AC_c_in'] # Minimum input power of the PV2AC conversion pathway
    _tde = _t_CONSTANT > 0 # Binary variable to activate the first-order time delay element
    _ftde = 1 - np.exp(-_dt / _t_CONSTANT) # Factor of the first-order time delay element
    # Kann and dieser Stelle auf einen Verschiebung von tstart um 2 verzichtet werden. Dann fängt t bei 0 an
    # Was ,achen die Funktonen, die auf eine vorherigen STufe zugreifen?
    _tstart = np.maximum(2, 1 + _t_DEAD) # First time step with regard to the dead time of the system control

    _E_BAT *= 1000 # Capacity of the battery, conversion from W to kW

    _eta_BAT /= 100
    
    # Check if the dead time can be ignored
    if _dt >= (3 * _t_CONSTANT):
        _tstart = 1
        T_DEAD = False
    else:
        T_DEAD = True

    # CHeck if the settling time can be ignored
    if _dt >= _t_DEAD + 3 * _t_CONSTANT:
        SETTLING = False
    else:
        SETTLING = True
    
    for t in range(_tstart - 1, _tend):
        
        # Energy content of the battery in the previous time step
        E_b0 = _soc0 * _E_BAT
        
        # Calculate the AC power of the battery system from the residual power
        # with regard to the dead time of the system control
        # Einen vorherigen Wert für Pr mitliefern, wenn die Totzeit berücksichtigt werden soll
        if T_DEAD:
            P_bs = _Pr[t - _t_DEAD]
        else:
            P_bs = _Pr[t]
        
        # Check if the battery holds enough unused capacity for charging or discharging
        # Estimated amount of energy in Ws that is supplied to or discharged from the storage unit.
        E_bs_est = P_bs * _dt / 1000
        
        if E_bs_est > 0 and E_bs_est > (_E_BAT - E_b0):
            P_bs = (_E_BAT - E_b0) / _dt
        # When charging take the correction factor into account
        elif E_bs_est < 0 and np.abs(E_bs_est) > (E_b0):
            P_bs = (E_b0) / _dt * (1-corr)
        
        # Adjust the AC power of the battery system due to the stationary 
        # deviations taking the minimum charging and discharging power into
        # account
        if P_bs > _P_AC2BAT_min:
            P_bs = np.maximum(_P_AC2BAT_min, P_bs + _P_AC2BAT_DEV)
            
        elif P_bs < -_P_BAT2AC_min:
            P_bs = np.minimum(-_P_BAT2AC_min, P_bs - _P_BAT2AC_DEV)
            
        else:
            P_bs = 0
        
        # Limit the AC power of the battery system to the rated power of the
        # battery converter
        P_bs = np.maximum(-_P_BAT2AC_out * 1000, np.minimum(_P_AC2BAT_in * 1000, P_bs))

        # Adjust the AC power of the battery system due to the settling time
        # (modeled by a first-order time delay element) Hier hat der Schritt vorher eine Null?
        # Muss der vorherige Wert mit übergeben werden?
        if SETTLING:
            P_bs = _tde * _Pbs[t-1] + _tde * (P_bs - _Pbs[t-1]) * _ftde + P_bs * (not _tde)
        
        # Decision if the battery should be charged or discharged
        if P_bs > 0 and _soc0 < 1 - _th * (1 - _SOC_h):
            # The last term th*(1-SOC_h) avoids the alternation between
            # charging and standby mode due to the DC power consumption of the
            # battery converter when the battery is fully charged. The battery
            # will not be recharged until the SOC falls below the SOC-threshold
            # (SOC_h) for recharging from PV.
            
            # Normalized AC power of the battery system
            p_bs = P_bs / _P_AC2BAT_in / 1000
            
            # DC power of the battery affected by the AC2BAT conversion losses
            # of the battery converter
            P_bat = np.maximum(0, P_bs - (_AC2BAT_a_in * p_bs * p_bs + _AC2BAT_b_in * p_bs + _AC2BAT_c_in))
            
        elif P_bs < 0 and _soc0 > 0:
            
            # Normalized AC power of the battery system
            p_bs = np.abs(P_bs / _P_BAT2AC_out / 1000)
            
            # DC power of the battery affected by the BAT2AC conversion losses
            # of the battery converter
            P_bat = P_bs - (_BAT2AC_a_out * p_bs * p_bs + _BAT2AC_b_out * p_bs + _BAT2AC_c_out)

        else: # Neither charging nor discharging of the battery
            
            # Set the DC power of the battery to zero
            P_bat = 0
        
        # Decision if the standby mode is active
        if P_bat == 0 and _soc0 <= 0: # Standby mode in discharged state
            
            # DC and AC power consumption of the battery converter
            P_bat = -np.maximum(0, _P_SYS_SOC0_DC)
            P_bs = _P_SYS_SOC0_AC
                
        elif P_bat == 0 and _soc0 > 0: # Standby mode in fully charged state
            
            # DC and AC power consumption of the battery converter
            P_bat = -np.maximum(0, _P_SYS_SOC1_DC)
            P_bs = _P_SYS_SOC1_AC
    
        # Transfer the realized AC power of the battery system and 
        # the DC power of the battery
        _Pbs[t] = P_bs
        _Pbat[t] = P_bat
        
        # Change the energy content of the battery Wx to Wh conversion
        if P_bat > 0:
            E_b = E_b0 + P_bat * np.sqrt(_eta_BAT) * _dt / 3600
        
        elif P_bat < 0:
            E_b = E_b0 + P_bat / np.sqrt(_eta_BAT) * _dt / 3600
        
        else:
            E_b = E_b0
        
        # Calculate the state of charge of the battery
        _soc0 = E_b / (_E_BAT)
        _soc[t] = _soc0
        
        # Adjust the hysteresis threshold to avoid alternation
        # between charging and standby mode due to the DC power
        # consumption of the battery converter.
        if _th and _soc[t] > _SOC_h or _soc[t] > 1:
            _th = True
        else:
            _th = False 

    return _Pbat, _Pbs, _soc, _soc0

def run_loss_PV(_E_BAT, _P_PV2AC_in, _P_PV2AC_out, _P_PV2BAT_in, _P_BAT2PV_out, _PV2AC_a_in, _PV2AC_b_in, _PV2AC_c_in, _PV2BAT_a_in, _PV2BAT_b_in, _PV2BAT_c_in, _PV2AC_a_out, _PV2AC_b_out, _PV2AC_c_out, _BAT2PV_a_out, _BAT2PV_b_out, _BAT2PV_c_out, _eta_BAT, _SOC_h, _P_PV2BAT_DEV, _P_BAT2AC_DEV, _P_SYS_SOC1_DC, _P_SYS_SOC0_AC, _P_SYS_SOC0_DC, _tend, _soc0, _Pac, _Ppv, _Ppv2bat_in, _Ppv2ac_out, _Pbat2pv_out, _Ppvbs, _Pbat, _soc, _dt, _th, _t_DEAD, _t_CONSTANT):
    # Initialization of particular variables
    P_PV2BAT_min = _PV2BAT_c_in # Minimum DC charging power
    P_BAT2PV_min = _BAT2PV_c_out # Minimum DC discharging power

    # Initialization of particular variables
    _tde = _t_CONSTANT > 0 # Binary variable to activate the first-order time delay element
    _ftde = 1 - np.exp(-_dt / _t_CONSTANT) # Factor of the first-order time delay element
    _tstart = np.maximum(2, 1 + _t_DEAD) # First time step with regard to the dead time of the system control

    _E_BAT *= 1000 # Conversion from W to kW

    _eta_BAT /= 100
    # Ab hier beginnt die Schleife
    # Start of the time step simulation
    for t in range(_tstart - 1, _tend):
        
        # Energy content of the battery in the previous time step
        E_b0 = _soc0 * _E_BAT

        # Target AC output power of the PV-battery system to cover the AC power demand
        P_pvbs = np.minimum(_Pac(t - _t_DEAD), _P_PV2AC_out * 1000)
        
        # Normalized AC output power of the PV2AC conversion pathway
        ppv2ac = P_pvbs / _P_PV2AC_out / 1000
        
        # Target DC input power of the PV2AC conversion pathway
        P_pv2ac_in = P_pvbs + (_PV2AC_a_out * ppv2ac**2 + _PV2AC_b_out * ppv2ac + _PV2AC_c_out)

        # Residual power 
        P_rpv = _Ppv(t - _t_DEAD) - P_pv2ac_in
                
        # Decision if the battery should be charged or discharged
        if P_rpv > P_PV2BAT_min and _soc0 < 1 - _th * (1 - _SOC_h):
            '''
            The last term th*(1-SOC_h) avoids the alternation between
            charging and standby mode due to the DC power consumption of the
            battery converter when the battery is fully charged. The battery
            will not be recharged until the SOC falls below the SOC-threshold
            (SOC_h) for recharging from PV.
            '''
            # Charging power
            P_pv2bat_in = P_rpv
            
            # Adjust the charging power due to stationary deviations
            P_pv2bat_in = np.maximum(0, P_pv2bat_in + _P_PV2BAT_DEV)
            
            # Limit the charging power to the maximum charging power
            P_pv2bat_in = np.minimum(P_pv2bat_in, _P_PV2BAT_in * 1000)
            
            # Adjust the charging power due to the settling time
            # (modeled by a first-order time delay element)
            P_pv2bat_in = _tde * _Ppv2bat_in[t-1] + _tde * (P_pv2bat_in - _Ppv2bat_in[t-1]) * _ftde + P_pv2bat_in * (not _tde) 
            
            # Limit the charging power to the current power output of the PV generator
            P_pv2bat_in = np.minimum(P_pv2bat_in, _Ppv[t])
            
            # Normalized charging power
            ppv2bat = P_pv2bat_in / _P_PV2BAT_in / 1000
            
            # DC power of the battery
            P_bat = np.maximum(0, P_pv2bat_in - (_PV2BAT_a_in * ppv2bat**2 + _PV2BAT_b_in * ppv2bat + _PV2BAT_c_in))
                
            # Realized DC input power of the PV2AC conversion pathway
            P_pv2ac_in = _Ppv[t] - P_pv2bat_in
            
            # Limit the DC input power of the PV2AC conversion pathway
            P_pv2ac_in = np.minimum(P_pv2ac_in, _P_PV2AC_in * 1000)
            
            # Recalculate Ppv(t) with limited PV2AC input power
            _Ppv[t] = P_pv2ac_in + P_pv2bat_in
            
            # Normalized DC input power of the PV2AC conversion pathway
            ppv2ac = P_pv2ac_in / _P_PV2AC_in / 1000
            
            # Realized AC power of the PV-battery system
            P_pv2ac_out = np.maximum(0, P_pv2ac_in - (_PV2AC_a_in * ppv2ac**2 + _PV2AC_b_in * ppv2ac + _PV2AC_c_in))
            P_pvbs = P_pv2ac_out
            
            # Transfer the final values
            _Ppv2ac_out[t] = P_pv2ac_out
            _Ppv2bat_in[t] = P_pv2bat_in
            
        elif P_rpv < -P_BAT2PV_min and _soc0 > 0:
            # Target discharging power of the battery
            P_bat2pv_out = np.abs(P_rpv)
            
            # Adjust the discharging power due to the stationary deviations
            P_bat2ac_out = np.maximum(0, P_bat2ac_out + _P_BAT2AC_DEV)
            
            # Adjust the discharging power to the maximum discharging power
            P_bat2ac_out = np.minimum(P_bat2pv_out, _P_BAT2PV_out * 1000)
            
            # Adjust the discharging power due to the settling time
            # (modeled by a first-order time delay element)
            P_bat2ac_out = _tde * _Pbat2pv_out[t-1] + _tde * (P_bat2ac_out - _Pbat2pv_out[t-1]) * _ftde + P_bat2ac_out * (not _tde)
            
            # Recalculate   np. withimum limited PV2AC input power
            _Ppv[t] = np.minimum(_P_PV2AC_in * 1000, _Ppv[t])
            
            # Limit the discharging power to the maximum AC power output of the PV-battery system
            P_bat2ac_out = np.minimum(_P_PV2AC_in * 1000 - _Ppv[t], P_bat2ac_out)
            
            # Normalized discharging power
            pbat2pv = P_bat2ac_out / _P_BAT2PV_out / 1000
            
            # DC power of the battery affected by the BAT2PV conversion losses
            P_bat = -1 * (P_bat2ac_out + (_BAT2PV_a_out * pbat2pv**2 + _BAT2PV_b_out * pbat2pv + _BAT2PV_c_out))
            
            # Realized DC input power of the PV2AC conversion pathway
            P_pv2ac_in = _Ppv[t] + P_bat2ac_out
            
            # Normalized DC input power of the PV2AC conversion pathway
            ppv2ac = P_pv2ac_in / _P_PV2AC_in / 1000
            
            # AC power of the PV-battery system
            P_pvbs = np.maximum(0, P_pv2ac_in - (_PV2AC_a_in * ppv2ac**2 + _PV2AC_b_in * ppv2ac + _PV2AC_c_in))
            P_pv2ac_out = P_pvbs
            
            # Transfer the final values
            _Ppv2ac_out[t] = P_pv2ac_out
            _Pbat2pv_out[t] = P_bat2ac_out
                
        else: # Neither charging nor discharging of the battery
            
            # Set the DC power of the battery to zero
            P_bat = 0
            
            # Limit the power output of the PV generator to the maximum input power
            # of the PV inverter
            _Ppv[t] = np.minimum(_Ppv[t], _P_PV2AC_in * 1000)
            
            # Normalized DC input power of the PV2AC conversion pathway
            ppv2ac = _Ppv[t] / _P_PV2AC_in / 1000
            
            # Realized AC power of the PV-battery system
            P_pvbs = np.maximum(0, _Ppv[t] - (_PV2AC_a_in * ppv2ac**2 + _PV2AC_b_in * ppv2ac + _PV2AC_c_in))

            # Transfer the final values
            _Ppv2ac_out[t] = P_pvbs
            
        # Decision if the standby mode is active
        if P_bat == 0 and _soc0 <= 0: # Standby mode in discharged state
            
            # DC power consumption of the battery converter
            P_bat = -np.maximum(0, _P_SYS_SOC0_DC)
            if P_pvbs == 0:        
                P_pvbs = -_P_SYS_SOC0_AC
                
        elif P_bat == 0 and P_pvbs > 0 and _soc0 > 0: # Standby mode in fully charged state

            # DC power consumption of the battery converter
            P_bat = -np.maximum(0, _P_SYS_SOC1_DC)
    

        # Transfer the realized AC power of the battery system and 
        # the DC power of the battery
        _Ppvbs[t] = P_pvbs
        _Pbat[t] = P_bat
        
        # Change the energy content of the battery Wx to Wh conversio
        if P_bat > 0:
            E_b = E_b0 + P_bat * np.sqrt(_eta_BAT) * _dt / 3600
        
        elif P_bat < 0:
            E_b = E_b0 + P_bat / np.sqrt(_eta_BAT) * _dt / 3600
        
        else:
            E_b = E_b0
        
        # Calculate the state of charge of the battery
        _soc0 = E_b / (_E_BAT)
        _soc[t] = _soc0
        
        # Adjust the hysteresis threshold to avoid alternation
        # between charging and standby mode due to the DC power
        # consumption of the battery converter.
        if _th and _soc[t] > _SOC_h or _soc[t] > 1:
            _th = True
        else:
            _th = False 

    return _soc, _soc0, _Ppv, _Ppvbs, _Pbat, _Ppv2ac_out, _Pbat2pv_out, _Ppv2bat_in

def bat_res_mod(_parameter, _Pl, _Ppv, _Pbat, _dt, *args):
    '''
    TODO
    Bei der Umwandlung zu MWh dt brücksichtigen!
    '''    
    
    _E = dict()

    if _parameter['Top'] == 'AC': # AC-coupled systems
        
        _Ppvs = args[0] # AC output power of the PV system
        _Pbs = args[1] # AC power of the battery system
        _Pperi = args[2] # Additional power consumption of the other system components
        
    elif _parameter['Top'] == 'DC' or _parameter['Top'] == 'PV': # DC- and PV-coupled systems
        
        _Ppv2ac = args[0] # AC output power of the PV2AC conversion pathway
        _Ppv2bat_in = args[1] # Input power of the PV2BAT conversion pathway
        _Ppvbs = args[2] # AC power of the PV-battery system
        _Pperi = args[3] # Additional power consumption of the other system components

        _Ppv2ac_in = _Ppv - _Ppv2bat_in # Input power of the PV2AC conversion pathway
        
    # Total load including the power consumption of the other system components
    _Plt = _Pl + _Pperi
    # DC input power of the battery (charged)
    _Pbatin = np.maximum(0, _Pbat)
    # DC output power of the battery (discharged)
    _Pbatout = np.minimum(0, _Pbat)
    # Maximum PV feed-in power
    _P_ac2g_max = _parameter['p_ac2g_max'] * _parameter['P_PV'] * 1000

    if _parameter['Top'] == 'AC': # AC-coupled systems
        
        # Residual power without curtailment
        _Pr = _Ppvs - _Plt
        # AC input power of the battery system
        _Pac2bs = np.maximum(0, _Pbs)
        # AC output power of the battery system
        _Pbs2ac = np.minimum(0, _Pbs)
        # Negative residual power (residual load demand)
        _Prn = np.minimum(0, _Pr)
        # Positive residual power (surplus PV power)
        _Prp = np.maximum(0, _Pr)
        # Direct use of PV power by the load 
        _Ppvs2l = np.minimum(_Ppvs, _Plt)
        # PV charging power
        _Ppvs2bs = np.minimum(_Prp, _Pac2bs)
        # Grid charging power
        _Pg2bs = np.maximum(_Pac2bs - _Prp, 0)
        # Grid supply power of the load
        _Pg2l = np.minimum(_Prn - _Pbs2ac, 0)
        # Battery supply power of the load
        _Pbs2l = np.maximum(_Prn, _Pbs2ac)
        # Battery feed-in power
        _Pbs2g = np.minimum(_Pbs2ac - _Prn, 0)
        # PV feed-in power including curtailment 
        _Ppvs2g = np.minimum(np.maximum(_Prp - _Pac2bs, 0), _P_ac2g_max)
        # Power demand from the grid
        _Pg2ac = _Pg2l - _Pg2bs
        # Feed-in power to the grid
        _Pac2g = _Ppvs2g - _Pbs2g
        # Grid power
        _Pg = _Pac2g + _Pg2ac
        # Curtailed PV power (AC output power)
        _Pct = np.maximum(_Prp - _Pac2bs, 0) - _Ppvs2g
        # AC output power of the PV system including curtailment
        _Ppvs = _Ppvs - _Pct
        # Residual power including curtailment
        _Pr = _Ppvs - _Plt
        # Index for PV curtailment
        _idx = np.where(_Pct > 0)[0]
        
        for i in range(len(_idx)):
            
            _tct = _idx[i]
            # Normalized output power of the PV inverter
            _ppvinvout = _Ppvs[_tct] / _parameter['P_PV2AC_out'] / 1000
            # DC output power of the PV generator taking into account the
            # conversion and curtailment losses
            _Ppv[_tct] = _Ppvs[_tct] + (_parameter['PV2AC_a_out'] * _ppvinvout**2 + _parameter['PV2AC_b_out'] * _ppvinvout + _parameter['PV2AC_c_out'])           

    elif _parameter['Top'] == 'DC' or _parameter['Top'] == 'PV': # DC- and PV-coupled systems
        
        # Grid power demand of the PV-battery system
        _Pg2pvbs = np.minimum(0, _Ppvbs)
        # AC input power of the PV-battery system
        _Pac2pvbs = _Pg2pvbs
        # AC output power of the PV-battery system
        _Ppvbs2ac = np.maximum(0, _Ppvbs)
        # Load supply power by the PV-battery system
        _Ppvbs2l = np.minimum(_Plt, _Ppvbs2ac)
        # Load supply power by the grid
        _Pg2l = _Plt - _Ppvbs2l
        # Direct use of PV power by the load 
        _Ppv2l = np.minimum(_Plt, _Ppv2ac)
        # PV feed-in power including curtailment 
        _Ppv2g = np.minimum(_Ppv2ac - _Ppv2l, _P_ac2g_max)
        # Curtailed PV power (AC output power)
        _Pct = _Ppv2ac - _Ppv2l - _Ppv2g
        
        if np.sum(_Pct) > 0:            
            # Power of the PV-battery system including curtailment
            _Ppvbs = _Ppvbs - _Pct
            # AC output power of the PV-battery system including curtailment
            _Ppvbs2ac = np.maximum(0, _Ppvbs)
            # AC output power of the PV2AC conversion pathway including curtailment
            _Ppv2ac = _Ppv2ac - _Pct
            # Index for PV curtailment
            _idx = np.where(_Pct > 0)[0]
                            
            for i in range(len(_idx)):
                
                _tct = _idx[i]
                # Specific AC output power of the PV2AC conversion pathway
                _ppv2ac = _Ppv2ac[_tct] / _parameter['P_PV2AC_out'] / 1000
                # DC input power of the PV2AC conversion pathway including curtailment
                _Ppv2ac_in[_tct] = _Ppv2ac[_tct] + (_parameter['PV2AC_a_out'] * _ppv2ac **2 + _parameter['PV2AC_b_out'] * _ppv2ac + _parameter['PV2AC_c_out'])
            
            # DC output power of the PV generator including curtailment
            _Ppv = _Ppv2ac_in + _Ppv2bat_in
            
        # Grid power including curtailment
        _Pg = _Ppvbs-_Plt
        # Feed-in power to the grid including curtailment
        _Pac2g = np.maximum(0, _Pg)
        # Power demand from the grid
        _Pg2ac = np.minimum(0, _Pg)
    
    # 2. Energy sums
    # All variables in MWh
        
    # Electrical demand including the energy consumption of the other system components
    _E['El'] = np.sum(np.abs(_Plt)) * _dt  / 3.6e9
    # DC output of the PV generator including curtailment
    _E['Epv'] = np.sum(np.abs(_Ppv)) * _dt / 3.6e9
    # DC input of the battery (charged)
    _E['Ebatin'] = np.sum(np.abs(_Pbatin)) * _dt / 3.6e9
    # DC output of the battery (discharged)
    _E['Ebatout'] = np.sum(np.abs(_Pbatout)) * _dt / 3.6e9
    # Grid feed-in
    _E['Eac2g'] = np.sum(np.abs(_Pac2g)) * _dt / 3.6e9
    # Grid demand
    _E['Eg2ac'] = np.sum(np.abs(_Pg2ac)) * _dt / 3.6e9
    # Load supply by the grid
    _E['Eg2l'] = np.sum(np.abs(_Pg2l)) * _dt / 3.6e9
    # Demand of the other system components
    _E['Eperi'] = np.sum(np.abs(_Pperi)) * _dt / 3.6e9
    # Curtailed PV energy
    _E['Ect'] = np.sum(np.abs(_Pct)) * _dt / 3.6e9
    
    if _parameter['Top'] == 'AC': # AC-coupled systems
        
        # AC output of the PV system including curtailment
        _E['Epvs'] = np.sum(np.abs(_Ppvs)) * _dt / 3.6e9
        # AC input of the battery system
        _E['Eac2bs'] = np.sum(np.abs(_Pac2bs)) * _dt / 3.6e9
        # AC output of the battery system
        _E['Ebs2ac'] = np.sum(np.abs(_Pbs2ac)) * _dt / 3.6e9
        # Direct use of PV energy
        _E['Epvs2l'] = np.sum(np.abs(_Ppvs2l)) * _dt / 3.6e9
        # PV charging
        _E['Epvs2bs'] = np.sum(np.abs(_Ppvs2bs)) * _dt / 3.6e9
        # Grid charging
        _E['Eg2bs'] = np.sum(np.abs(_Pg2bs)) * _dt / 3.6e9
        # PV feed-in
        _E['Epvs2g'] = np.sum(np.abs(_Ppvs2g)) * _dt / 3.6e9
        # Load supply by the battery system
        _E['Ebs2l'] = np.sum(np.abs(_Pbs2l)) * _dt / 3.6e9
        # Battery feed-in
        _E['Ebs2g'] = np.sum(np.abs(_Pbs2g)) * _dt / 3.6e9
        
    elif _parameter['Top'] == 'DC' or _parameter['Top'] == 'PV': # DC- and PV-coupled systems
        
        # Grid demand of the PV-battery system
        _E['Eg2pvbs'] = np.sum(np.abs(_Pg2pvbs)) * _dt / 3.6e9
        # AC input of the PV-battery system
        _E['Eac2pvbs'] = np.sum(np.abs(_Pac2pvbs)) * _dt / 3.6e9
        # AC output of the PV-battery system
        _E['Epvbs2ac'] = np.sum(np.abs(_Ppvbs2ac)) * _dt / 3.6e9
        # Load supply by the PV-battery system
        _E['Epvbs2l'] = np.sum(np.abs(_Ppvbs2l)) * _dt / 3.6e9  
    
    return _E        
